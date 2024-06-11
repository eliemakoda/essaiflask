from flask import Flask, render_template, send_from_directory, request, Response, make_response
import pandas as pd
import geopandas as gpd
import os
from fpdf import FPDF
from flask import Flask, render_template, request, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from flask import jsonify
from flask import flash, redirect
from pandas import json_normalize
import numpy as np  # Import the NumPy library for NaN handling
import json
from sqlalchemy.orm.state import InstanceState  # Import InstanceState
from sqlalchemy.inspection import inspect
import logging
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.utils import secure_filename
import os
from flask import session
from datetime import date
import matplotlib.pyplot as plt
import matplotlib
import matplotlib.patches as mpatches
from flask_bcrypt import check_password_hash
from flask_bcrypt import Bcrypt
from sqlalchemy import desc
import openpyxl
from flask import send_file
from flask_mail import Mail, Message
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart  # Add this line
from flask_wtf import FlaskForm
import mplcursors
import io
import seaborn as sns

from sklearn.ensemble import RandomForestClassifier, BaggingClassifier
from sklearn.model_selection import train_test_split, RandomizedSearchCV, KFold, RepeatedStratifiedKFold, cross_val_score
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix,ConfusionMatrixDisplay,accuracy_score, precision_score, recall_score, f1_score
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.naive_bayes import MultinomialNB

from xgboost import XGBClassifier
from lightgbm import LGBMClassifier
from catboost import CatBoostClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from time import time
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import BaggingRegressor, RandomForestRegressor
from sklearn.decomposition import PCA
from sklearn import metrics
from sklearn.metrics import mean_squared_error 
from scipy.stats import randint

app = Flask(__name__)
bcrypt = Bcrypt()
# Set the secret key
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 'luigitimola@gmail.com'
app.config['MAIL_PASSWORD'] = 'yvmw udfs zijy iprs'

mail = Mail(app)

# Configure your MySQL database connection (replace with your actual credentials)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+mysqlconnector://root:@localhost/taallakedb'#dbname at the end of url
app.config['SQLALCHEMY_ECHO'] = True  # Enable query logging
db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'  # Set the login view
app.config['SECRET_KEY'] = 'your_secret_key'
matplotlib.use('Agg')
class WaterQuality(db.Model):
    __tablename__ = 'taaltb'  # Specify the custom table name if needed
    id = db.Column(db.Integer, primary_key=True)
    stationid = db.Column(db.String(50), name="stationid")
    Barangay = db.Column(db.String(50), name="Barangay")
    Month = db.Column(db.String(50), name="Month")
    Year = db.Column(db.String(50), name="Year")
    pH = db.Column(db.String(50), name="pH")
    Ammonia = db.Column(db.String(50), name="Ammonia")
    DO = db.Column(db.String(50), name="DO")
    Nitrate = db.Column(db.String(50), name="Nitrate")
    Phosphate = db.Column(db.String(50), name="Phosphate")
    Time = db.Column(db.String(50), name="Time")  # New field for Time
    WeatherCondition = db.Column(db.String(50), name="weater-condition")  # New field for Weather Condition
    WindDirection = db.Column(db.String(50), name="wind-direction")  # New field for Wind Direction
    ColorOfWater = db.Column(db.String(50), name="color-of-water")  # New field for Color of Water
    AirTemperature = db.Column(db.String(50), name="air-temperature")  # New field for Air Temperature
    WaterTransparency = db.Column(db.String(50), name="water-transparency")  # New field for Water Transparency
    WaterTemperature = db.Column(db.String(50), name="water-temperature")  # New field for Water Temperature



class User(UserMixin, db.Model):
    __tablename__ = 'usertb'
    id = db.Column(db.Integer, primary_key=True)
    fname = db.Column(db.String(50), nullable=False)
    mname = db.Column(db.String(50), nullable=False)
    lname = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(50), nullable=False)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(50), nullable=False)
    userType = db.Column(db.String(50), nullable=False)
    status = db.Column(db.String(50), nullable=False)
    profile_image = db.Column(db.String(255), nullable=True)  # Assuming a maximum file path length of 255 characters

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ... (your WaterQuality class and other routes)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()

        if user and check_password_hash(user.password, password):
            if user.status != 'disabled':
                # Log the user in using Flask-Login
                login_user(user)

                # Store user-related information in the session
                session['user_id'] = user.id
                session['fname'] = user.fname
                session['lname'] = user.lname
                session['userType'] = user.userType
                session['status'] = user.status
                session['profile_image'] = user.profile_image

                # Print user information for debugging
                print(f"User ID: {user.id}")
                print(f"First Name: {user.fname}")
                print(f"Last Name: {user.lname}")
                print(f"Profile Image: {user.userType}")

                return redirect(url_for('adminpanel'))#change to adminpanel
            else:
                flash('Login failed. User is disabled.', 'error')
        else:
            flash('Login failed. Check your username and password.', 'error')

    # Redirect to the index page if login fails
    return render_template('login.html')



#=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
# Admin Panel
@app.route('/adminpanel')
def adminpanel():
    # Render the admin panel
    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    userType = session.get('userType')
    profile_image = session.get('profile_image')

    # Load data from the database
    db_data = WaterQuality.query.all()
    # Convert the database data to a list of dictionaries
    db_data_dict = []
    for record in db_data:
        # Convert the SQLAlchemy model object to a dictionary
        record_dict = {
            'stationid': record.stationid,
            'Barangay': record.Barangay,
            'Month': record.Month,
            # 'Year': int(record.Year) if record.Year is not None else None,  # Check for None
            # 'pH': float(record.pH) if record.pH is not None else None,  # Convert to float, and check for None
            # 'Ammonia': float(record.Ammonia) if record.Ammonia is not None else None,  # Convert to float, and check for None
            # 'DO': float(record.DO) if record.DO is not None else None,  # Convert to float, and check for None
            # 'Nitrate': float(record.Nitrate) if record.Nitrate is not None else None,  # Convert to float, and check for None
            # 'Phosphate': float(record.Phosphate) if record.Phosphate is not None else None  # Convert to float, and check for None
            'Year': int(record.Year) if record.Year.strip() else None,  # Check for empty string
            'pH': float(record.pH) if record.pH.strip() else None,  # Convert to float, and check for empty string
            'Ammonia': float(record.Ammonia) if record.Ammonia.strip() else None,  # Convert to float, and check for empty string
            'DO': float(record.DO) if record.DO.strip() else None,  # Convert to float, and check for empty string
            'Nitrate': float(record.Nitrate) if record.Nitrate.strip() else None,  # Convert to float, and check for empty string
            'Phosphate': float(record.Phosphate) if record.Phosphate.strip() else None  # Convert to float, and check for empty string

        }

        db_data_dict.append(record_dict)

    # Convert the list of dictionaries to a Pandas DataFrame
    db_data_df = pd.DataFrame(db_data_dict)

    # Add more conversions as needed

    # Load data from the CSV file
    csv_data = pd.read_csv('taaldata.csv', encoding='ISO-8859-1')
    # Standardize the CSV data
    numerical_columns = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']
    for col in numerical_columns:
        csv_data[col] = csv_data[col].astype(float)

    # Calculate WQI for the CSV data and the database data
    db_data_df = calculate_wqi(db_data_df)
    csv_data = calculate_wqi(csv_data)

    for data in [db_data_df, csv_data]:
        for col in numerical_columns:
            data[col] = data[col].apply(lambda x: None if pd.isna(x) else x)

    # Combine the data from the database and CSV
    combined_data_df = pd.concat([db_data_df, csv_data], ignore_index=True)

    # Convert the combined data DataFrame to JSON format
    combined_data_json = combined_data_df.to_json(orient='records')


    lake_df = pd.read_csv('taaldata.csv',encoding='ISO-8859-1')
    lake_df = calculate_wqi(lake_df)  # Calculate WQI
    lake_csv_data_with_wqi_json = lake_df.to_json(orient='records')


    filtered_data = combined_data_df[(combined_data_df['Month'] == 'January ') & (combined_data_df['Year'] == 2019)]
    print('dao chart')
    print(filtered_data)
    # Get selected month and year from the request query parameters
    selected_month = 'January'
    selected_year = 2019

    # Filter data based on selected month and year
    # filtered_data = combined_data_df[(combined_data_df['Month'] == selected_month) & (combined_data_df['Year'] == int(selected_year))]#

    # Extract barangays and parameter values
    barangays = filtered_data['Barangay'].unique()
    parameters = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']

    parameter_values = {param: [] for param in parameters}
    parameter_suitability = {param: [] for param in parameters}

    for barangay in barangays:
        barangay_data = filtered_data[filtered_data['Barangay'] == barangay]
        for param in parameters:
            values = barangay_data[param].values
            suitability = [check_suitability(value, param)['suitability'] for value in values]
            parameter_values[param].append(values.mean())
            parameter_suitability[param].append(suitability[0])

    # Plot the data
    plt.figure(figsize=(12, 6))

    bar_width = 0.15
    index = np.arange(len(barangays))

    # Plot bars for each parameter
    for i, param in enumerate(parameters):
        offset = bar_width * (i - len(parameters) // 2)
        colors = ['blue' if s == 'Suitable' else 'red' for s in parameter_suitability[param]]
        bars = plt.bar(index + offset, parameter_values[param], bar_width, label=param, color=colors)
        for bar in bars:
            height = bar.get_height()
            plt.annotate(f'{param}', xy=(bar.get_x() + bar.get_width() / 2, height),
                         xytext=(0, 3),  # 3 points vertical offset
                         textcoords="offset points",
                         ha='center', va='bottom', rotation=90)  # Rotate label by 90 degrees
    green_patch = mpatches.Patch(color='blue', label='Suitable')
    red_patch = mpatches.Patch(color='red', label='Not Suitable')
    plt.legend(handles=[green_patch, red_patch])
    plt.xlabel('Barangay')
    plt.ylabel('Value')
    plt.title(f'Water Quality Parameters for {selected_month} {selected_year}')
    plt.xticks(index, barangays, rotation=45, ha='right')
    plt.tight_layout()

    # Save the plot as an image
    plot_path = f'static/bar_chart_1_{selected_month}_{selected_year}.png'
    plt.savefig(plot_path)
    plt.close()


    # Close the plot to free up resources
    plt.close()
    
    #=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-==-=-=-=-
         # Create a DataFrame from the sample data
    df = pd.DataFrame(combined_data_df)

    # Filter the data for January 2018
    january_2018_data = df[(df['Month'] == 'January') & (df['Year'] == 2018)]

    # Iterate over each row and check suitability
    suitability_data = []
    for index, row in january_2018_data.iterrows():
        station_id = row['stationid']
        barangay = row['Barangay']
        month = row['Month']
        year = row['Year']
        suitability = {}
        for parameter in ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']:
            value = row[parameter]
            suitability[parameter] = check_suitability(value, parameter)
        suitability_data.append({'station_id': station_id, 'barangay': barangay, 'month': month, 'year': year, 'suitability': suitability})

    # Extract barangays and parameter values
    barangays = january_2018_data['Barangay'].unique()
    parameters = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']

    parameter_values = {param: [] for param in parameters}
    parameter_suitability = {param: [] for param in parameters}

    for barangay in barangays:
        barangay_data = january_2018_data[january_2018_data['Barangay'] == barangay]
        for param in parameters:
            values = barangay_data[param].values
            suitability = [check_suitability(value, param)['suitability'] for value in values]
            parameter_values[param].append(values.mean())
            parameter_suitability[param].append(suitability[0])

    # Plot the data
    plt.figure(figsize=(12, 6))

    bar_width = 0.15
    index = np.arange(len(barangays))

    # Plot bars for each parameter
    # Plot bars for each parameter
    for i, param in enumerate(parameters):
        offset = bar_width * (i - len(parameters) // 2)
        colors = ['blue' if s == 'Suitable' else 'red' for s in parameter_suitability[param]]
        bars = plt.bar(index + offset, parameter_values[param], bar_width, label=param, color=colors)
        for bar in bars:
            height = bar.get_height()
            plt.annotate(f'{param}', xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='bottom', rotation=90)  # Rotate label by 90 degrees
    #no hover
    plt.xlabel('Barangay')
    plt.ylabel('Value')
    plt.title('Water Quality Parameters by Barangay (January 2018)')
    plt.xticks(index, barangays, rotation=45, ha='right')
    #plt.legend()
    plt.tight_layout()

    # Add tooltips using mplcursors
    mplcursors.cursor(hover=True).connect('add', lambda sel: sel.annotation.set_text(sel.artist.get_label()))

    # Save the plot as an image
    plot_path1 = 'static/barangay_parameter_plot.png'
    plt.savefig(plot_path1)
    plt.close()

    # Convert the combined data DataFrame to JSON format
    combined_data_json = combined_data_df.to_json(orient='records')
    return render_template('adminpanel.html', user_id=user_id, fname=fname, lname=lname, userType=userType, profile_image=profile_image,plot_path=plot_path,plot_path1=plot_path1)

# Admin Panel
@app.route('/user')
def user():
     # Load data from the database
    db_data = User.query.all()

    # Convert the database data to a list of dictionaries
    db_data_dict = [
        {
            'id': record.id,
            'fname': record.fname,
            'mname': record.mname,
            'lname': record.lname,
            'email': record.email,
            'username': record.username,
            'password': record.password,
            'userType' :record.userType,
            'status' :record.status
        }
        for record in db_data
    ]

    # Print the data for debugging
    #print('DB Data:', db_data_dict)

    # Convert the list of dictionaries to a JSON string
    user_data_json = json.dumps(db_data_dict)
    # Render the admin panel
    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    userType = session.get('userType')
    profile_image = session.get('profile_image')
    
    return render_template('user.html',user_data=user_data_json, user_id=user_id, fname=fname, lname=lname, userType=userType, profile_image=profile_image)

# =-=-=-==-=-=-=-=-=-=-=-==-=-==-==--==-=-=-=-=--=-=-=-=-=-=-=-=---=-=-==-=-=-=-=-=-=-=-=-=-=-=-=--=-=-=-=-=-=-=--=-=-=-=-===-=-==-==-==-=-=-
@app.route('/update_status', methods=['POST'])
def update_status():
    try:
        if request.method == 'POST':
            # Retrieve data from the form
            id = request.form.get('id')
            status = request.form.get('status')


            # Find the user by ID
            user_to_update = User.query.get(id)

            # Update user data if found
            if user_to_update:
                user_to_update.status = 'disabled'

                # Commit the changes to the database
                db.session.commit()

                # Flash message for success
                flash('User disabled successfully', 'success')

                # Redirect to usermanagement or another page
                return redirect(url_for('user'))

            else:
                flash('User not found', 'error')
                return redirect(url_for('user'))  # Redirect with an error flash message

    except Exception as e:
        flash('Error occurred while updating user', 'error')
        return redirect(url_for('user'))  # Redirect with an error flash message
    

@app.route('/update_status1', methods=['POST'])
def update_status1():
    try:
        if request.method == 'POST':
            # Retrieve data from the form
            id = request.form.get('id')
            status = request.form.get('status')


            # Find the user by ID
            user_to_update = User.query.get(id)

            # Update user data if found
            if user_to_update:
                user_to_update.status = 'active'

                # Commit the changes to the database
                db.session.commit()

                # Flash message for success
                flash('User disabled successfully', 'success')

                # Redirect to usermanagement or another page
                return redirect(url_for('user'))

            else:
                flash('User not found', 'error')
                return redirect(url_for('user'))  # Redirect with an error flash message

    except Exception as e:
        flash('Error occurred while updating user', 'error')
        return redirect(url_for('user'))  # Redirect with an error flash message

# # Announcment
# @app.route('/activities')
# def activities():
#     # Render the admin panel
#     return render_template('activities.html')
# Announcment
 # Read the CSV file into a DataFrame
@app.route('/charts_dao')
def charts_dao():
    # Read the CSV file into a DataFrame
    df = pd.read_csv('taaldata.csv',encoding='ISO-8859-1')

    # Define the function to check suitability of each parameter
    def check_suitability(value, parameter):
        suitability = ''
        if parameter == 'pH':
            suitability = 'Suitable' if 6.5 <= value <= 9.0 else 'Not Suitable'
        elif parameter == 'Ammonia':
            suitability = 'Suitable' if value < 0.05 else 'Not Suitable'
        elif parameter == 'DO':
            suitability = 'Suitable' if value > 5.0 else 'Not Suitable'
        elif parameter == 'Nitrate':
            suitability = 'Suitable' if value < 7.0 else 'Not Suitable'
        elif parameter == 'Phosphate':
            suitability = 'Suitable' if value < 0.5 else 'Not Suitable'
        return {'value': value if pd.notnull(value) else 'NaN', 'suitability': suitability}

    # Iterate over each row and check suitability
    suitability_data = []
    for index, row in df.iterrows():
        station_id = row['stationid']
        barangay = row['Barangay']
        month = row['Month']
        year = row['Year']
        suitability = {}
        for parameter in ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']:
            value = row[parameter]
            suitability[parameter] = check_suitability(value, parameter)
        suitability_data.append({'station_id': station_id, 'barangay': barangay, 'month': month, 'year': year, 'suitability': suitability})

    # Get session data
    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    userType = session.get('userType')
    profile_image = session.get('profile_image')

    # Convert the data to JSON format
    suitability_data_json = json.dumps(suitability_data)

    # Render the template with the processed data and session data
    return render_template('charts_dao.html', data=suitability_data_json, user_id=user_id, fname=fname, lname=lname, userType=userType, profile_image=profile_image)

#
#
# Define the function to check suitability of each parameter
def check_suitability(value, parameter):
    suitability = ''
    if parameter == 'pH':
        suitability = 'Suitable' if 6.5 <= value <= 9.0 else 'Not Suitable'
    elif parameter == 'Ammonia':
        suitability = 'Suitable' if value < 0.05 else 'Not Suitable'
    elif parameter == 'DO':
        suitability = 'Suitable' if value > 5.0 else 'Not Suitable'
    elif parameter == 'Nitrate':
        suitability = 'Suitable' if value < 7.0 else 'Not Suitable'
    elif parameter == 'Phosphate':
        suitability = 'Suitable' if value < 0.5 else 'Not Suitable'
    return {'value': value if pd.notnull(value) else 'NaN', 'suitability': suitability}

class Contact(db.Model):
    __tablename__ = 'contacttb'  # Specify the custom table name if needed
    id = db.Column(db.Integer, primary_key=True)
    bldg_no = db.Column(db.String(50), name="bldg_no")
    brgy = db.Column(db.String(50), name="brgy")
    municipality = db.Column(db.String(50), name="municipality")
    province = db.Column(db.String(50), name="province")
    zip_code = db.Column(db.String(50), name="zip_code")
    mobile = db.Column(db.String(50), name="mobile")
    email = db.Column(db.String(50), name="email")
# Contacts
@app.route('/contacts')
def contacts():
     # Load data from the database
    db_data = Contact.query.all()
    # Convert the database data to a list of dictionaries
    db_data_dict = []
    for record in db_data:
        # Convert the SQLAlchemy model object to a dictionary
        record_dict = {
            'id':record.id,
            'bldg_no': record.bldg_no,
            'brgy': record.brgy,
            'municipality': record.municipality,
            'province': record.province,  # Check for None
            'zip_code': record.zip_code,  # Convert to float, and check for None
            'mobile': record.mobile,  # Convert to float, and check for None
            'email': record.email  # Convert to float, and check for None
        }

        db_data_dict.append(record_dict)

    # Convert the list of dictionaries to a Pandas DataFrame
    db_data_df = pd.DataFrame(db_data_dict)
    data_json = db_data_df.to_json(orient='records')
    # Render the admin panel
    # Render the admin panel
    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    userType = session.get('userType')
    profile_image = session.get('profile_image')
    
    return render_template('contacts.html',contact_json_data=data_json, user_id=user_id, fname=fname, lname=lname, userType=userType, profile_image=profile_image)


@app.route('/update_contact', methods=['POST'])
@login_required
def update_contact():
    try:
        # Retrieve data from the form
        id = request.form.get('id')
        bldg_no = request.form.get('udpatedBldg')
        brgy = request.form.get('updatedBrgy')
        municipality = request.form.get('updatedMunicipality')
        province = request.form.get('updatedProvince')
        zip_code = request.form.get('updatedZipcode')
        mobile = request.form.get('updatedMobile')
        email = request.form.get('updatedEmail')

        # Find the announcement by ID
        contact_to_update = Contact.query.get(id)

        # Check if the service exists
        if contact_to_update:
            # Update service data
            contact_to_update.bldg_no = bldg_no
            contact_to_update.brgy = brgy
            contact_to_update.municipality = municipality
            contact_to_update.province=province
            contact_to_update.zip_code=zip_code
            contact_to_update.mobile=mobile
            contact_to_update.email=email

            print("id:", id)
            print("bldg_no:", bldg_no)
            print("brgy:", brgy)
            print("municipality:", municipality)
            print("province:", province)
            print("zip_code:", zip_code)
            print("mobile:", mobile)
            print("email:", email)
            # Commit the changes to the database
            db.session.commit()

            # Flash message for success
            flash('Activity updated successfully', 'success')
        else:
            flash('Activity not found', 'error')

    except Exception as e:
        flash(f'Error occurred while updating contact: {e}', 'error')

    # Redirect to services page
    return redirect(url_for('contacts'))
#
#
#


#
#
#
UPLOAD_FOLDER = 'static/img'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

@app.route('/save_user', methods=['POST'])
def save_user():
    if request.method == 'POST':
        try:
            fname = request.form.get('fname')
            mname = request.form.get('mname')
            lname = request.form.get('lname')
            email = request.form.get('email')
            username = request.form.get('username')
            password = request.form.get('password')
            userType = request.form.get('userType')
            status = 'active'

            print(f"Received username: {username}")
            print(f"Received password: {password}")

            #hashed_password = generate_password_hash(password)
            hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')

            print(f"Hashed password: {hashed_password}")

            user_data = User(
                fname=fname,
                mname=mname,
                lname=lname,
                email=email,
                username=username,
                password=hashed_password,
                userType=userType,
                status=status
            )

            # Handle profile image upload
            if 'profile_image' in request.files:
                profile_image = request.files['profile_image']
                if profile_image.filename != '':
                    # Process the profile image as before
                    if '.' in profile_image.filename and \
                            profile_image.filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS:
                        filename = secure_filename(profile_image.filename)
                        filepath = os.path.join(UPLOAD_FOLDER, filename)
                        profile_image.save(filepath)
                        user_data.profile_image = filepath
                    else:
                        flash('Invalid file extension. Allowed extensions are png, jpg, jpeg, gif', 'error')
                        return redirect(request.url)
                else:
                    # No profile image provided, set profile_image to None
                    user_data.profile_image = None
            else:
                # No profile image provided, set profile_image to None
                user_data.profile_image = None

            db.session.add(user_data)
            db.session.commit()

            flash('User created successfully', 'success')

            print("User data saved successfully")

            return redirect(url_for('user'))
        except Exception as e:
            print(f"Error occurred: {e}")
            db.session.rollback()
            flash('Error occurred while saving data', 'error')
            return redirect(request.url)

    else:
        return jsonify({"message": "Invalid request method", "status": "error"})


@app.route('/update_user', methods=['POST'])
@login_required
def update_user():
    try:
        if request.method == 'POST':
            # Retrieve data from the form
            id = request.form.get('id')
            fname = request.form.get('updatedFname')
            mname = request.form.get('updatedMname')
            lname = request.form.get('updatedLname')
            email = request.form.get('updatedEmail')
            username = request.form.get('updatedUsername')
            password = request.form.get('updatedPassword')
            userType = request.form.get('updatedUserType')

            # Find the user by ID
            user_to_update = User.query.get(id)

            # Update user data if found
            if user_to_update:
                user_to_update.fname = fname
                user_to_update.mname = mname
                user_to_update.lname = lname
                user_to_update.email = email
                user_to_update.username = username
                user_to_update.password = password  # You may want to handle the password securely
                user_to_update.userType = userType

                # Commit the changes to the database
                db.session.commit()

                # Flash message for success
                flash('User updated successfully', 'success')

                # Redirect to usermanagement or another page
                return redirect(url_for('usermanagement'))

            else:
                flash('User not found', 'error')
                return redirect(url_for('usermanagement'))  # Redirect with an error flash message

    except Exception as e:
        flash('Error occurred while updating user', 'error')
        return redirect(url_for('usermanagement'))  # Redirect with an error flash message
#
#
#

# Abouts
@app.route('/abouts')
def abouts():
    # Render the admin panel
    return render_template('abouts.html')

#LOGIN
@app.route('/loginpage')
def loginpage():
    return redirect(url_for('login'))


import base64
import folium
from folium.plugins import HeatMap

# Define the coordinates for each Barangay
barangay_coordinates = {
    # 'Sampaloc': [13.960427097650134, 121.1145482751536],
    # 'Quiling': [13.9525959052652, 120.92609306647753],
    # 'San Isidro': [14.067116632318964, 120.9304048228642],
    # 'Buso-Buso': [14.145132074933167, 121.01922121441443],
    # 'Leviste': [14.107798507455575, 121.13049021559298],
    # 'Banaga': [14.145132074933167, 121.01922121441443],
    # 'Manalao': [14.107798507455575, 121.12049021559298],
    # 'Mataas na Kahoy': [13.960427097650134, 121.1145482751536],
    # 'Tanauan': [14.107798507455575, 121.12049021559298],
    # 'Agoncillo': [13.9525959052652, 120.92609306647753],
    # 'Laurel': [14.067116632318964, 120.9304048228642],
    # 'Talisay': [14.145132074933167, 121.01922121441443],
    # 'Nangkaan': [14.107798507455575, 121.12049021559298]
    'Tanauan': [14.0594166667, 121.056472222],
    'Nangkaan': [14.0027222222, 121.081111111],
    'Banaga': [14.0087222222, 120.957944444],
    'Manalaw': [13.9779444444, 120.962277778],
    'Leviste': [14.0686388889, 120.94525],
    'Buso-Buso': [14.0251944444, 120.955305556],
    'Quiling': [14.0869166667, 121.035305556],
    'Sampaloc': [14.0831666667, 120.9715]
}


import folium
import io
import base64

def create_heatmap(data, filename):
    # Define the coordinates for each Barangay
    barangay_coordinates = {
        'Tanauan': [14.0594166667, 121.056472222],
        'Nangkaan': [14.0027222222, 121.081111111],
        'Banaga': [14.0087222222, 120.957944444],
        'Manalaw': [13.9779444444, 120.962277778],
        'Leviste': [14.0686388889, 120.94525],
        'Buso-Buso': [14.0251944444, 120.955305556],
        'Quiling': [14.0869166667, 121.035305556],
        'Sampaloc': [14.0831666667, 120.9715]
    }

    # Create a folium map centered at the mean latitude and longitude of barangay coordinates
    map_center = [sum(coord[0] for coord in barangay_coordinates.values()) / len(barangay_coordinates), 
                  sum(coord[1] for coord in barangay_coordinates.values()) / len(barangay_coordinates)]
    m = folium.Map(location=map_center, zoom_start=10)
    
    # Parameters to include in the popup text 'pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate', 'wqi', 'wqc'
    parameters = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate', 'wqi', 'wqc']
    
    # Add heatmap layer for WQI with a custom color gradient
    for parameter in ['wqi']:
        heat_data = []
        
        # Iterate through the data and add to heat_data if the condition is met
        for _, row in data.iterrows():
            if row['Barangay'] in barangay_coordinates:
                weight = 0.1  # Set a default low weight value
                # Define weight based on the range of the 'wqi' parameter
                if 0 <= row[parameter] <= 50:
                    weight = 0.1  # Ensure a low weight for low 'wqi' values
                elif 51 <= row[parameter] <= 100:
                    weight = 0.7
                elif 101 <= row[parameter] <= 200:
                    weight = 0.85
                elif 201 <= row[parameter] <= 300:
                    weight = 0.93
                elif row[parameter] > 301:
                    weight = 1
                heat_data.append([barangay_coordinates[row['Barangay']][0], barangay_coordinates[row['Barangay']][1], weight])
        
        # Adjust radius and blur parameters to control merging effect
        folium.plugins.HeatMap(heat_data, 
                               name=parameter, 
                               min_opacity=0.5,  # Slightly reduce min_opacity to better visualize low weights
                               radius=25,       # Increase the radius to spread the points more
                               blur=15          # Adjust the blur to control merging effect
                              ).add_to(m)
        
        # Add pins for each barangay with all parameter values (including 'wqc') as popups
        for _, row in data.iterrows():
            if row['Barangay'] in barangay_coordinates:
                popup_text = f"<div style='width: 180px;' class = 'fs-3'><b>{row['Barangay']}</b><br>"
                for param in parameters:
                    popup_text += f"{param}: {row[param]}<br>"
                popup_text += "</div>"
                folium.Marker(location=barangay_coordinates[row['Barangay']], popup=popup_text).add_to(m)

    # Add legend for the color gradient
    legend_html = """
    <div style="
        position: fixed; 
        bottom: 50px; 
        left: 50px; 
        z-index:9999; 
        padding: 10px;
        background-color: white;
        border: 2px solid grey;
        font-size:14px;
        ">
        &nbsp; Heatmap Legend <br>
        &nbsp; <i class="fa fa-square" style="color:#0000FF"></i> 0 - 50 &nbsp; Excellent <br>
        &nbsp; <i class="fa fa-square" style="color:#00FF00"></i> 51 - 100 &nbsp; Good <br>
        &nbsp; <i class="fa fa-square" style="color:#FFFF00"></i> 101 - 200 &nbsp; Poor <br>
        &nbsp; <i class="fa fa-square" style="color:#FFA500"></i> 201 - 300 &nbsp; Very Poor <br>
        &nbsp; <i class="fa fa-square" style="color:#FF0000"></i> >= 301 &nbsp; Unsuitable <br>
    </div>
    """
    m.get_root().html.add_child(folium.Element(legend_html))
    
    # Add layer control to the map
    folium.LayerControl().add_to(m)
    
    # Save the map to a buffer
    buffer = io.BytesIO()
    m.save(buffer, close_file=False)
    buffer.seek(0)
    
    # Encode the buffer to base64 for HTML display
    map_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    # Save the map to a file
    m.save(filename)
    
    return map_base64

# GIS
@app.route('/gis')
def gis():
    # Render the admin panel
    #shapefile_path1 = r'shapefile/taalLake/Taal-Lake.shp'
    #data1 = gpd.read_file(shapefile_path1)
    #geojson_data1 = data1.to_json()

    lake_df = pd.read_csv('taaldata.csv',encoding='ISO-8859-1')
    lake_df = lake_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    lake_df = calculate_wqi(lake_df)  # Calculate WQI
    lake_csv_data_with_wqi_json = lake_df.to_json(orient='records')

    print('gis')
    print(lake_df)
    # Get month and year parameters from the request
    month = request.args.get('month')
    year = request.args.get('year')

    # Check if month and year are not None
    if month is not None and year is not None:
        # Convert month and year to integers
        month = month
        year = int(year)
    else:
        # Set default values to January 2019
        month = 'January'
        year = 2018

    # Filter lake_df based on month and year
    print(month)
    print(year)
    filtered_lake_df = lake_df[(lake_df['Month'] == month) & (lake_df['Year'] == year)]
    
    # Create the heatmap on the filtered data
    heatmap = create_heatmap(filtered_lake_df, 'heatmap.html')
    # Return statement as it is
    # Render the admin panel
    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    userType = session.get('userType')
    profile_image = session.get('profile_image')

    # Create the heatmap on the map
    #heatmap = create_heatmap(lake_df)
    #geojson_data1=geojson_data1, 
    return render_template('gis.html',  heatmap=heatmap, lake_csv_data=lake_csv_data_with_wqi_json, user_id=user_id, fname=fname, lname=lname, userType=userType, profile_image=profile_image)

@app.route('/gis1')
def gis1():
    # Render the admin panel
    #shapefile_path1 = r'shapefile/taalLake/Taal-Lake.shp'
    #data1 = gpd.read_file(shapefile_path1)
    #geojson_data1 = data1.to_json()

    lake_df = pd.read_csv('taaldata.csv',encoding='ISO-8859-1')
    lake_df = lake_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    lake_df = calculate_wqi(lake_df)  # Calculate WQI
    lake_csv_data_with_wqi_json = lake_df.to_json(orient='records')

    print('gis')
    print(lake_df)
    # Get month and year parameters from the request
    month = request.args.get('month')
    year = request.args.get('year')

    # Check if month and year are not None
    if month is not None and year is not None:
        # Convert month and year to integers
        month = month
        year = int(year)
    else:
        # Set default values to January 2019
        month = 'January'
        year = 2018

    # Filter lake_df based on month and year
    print(month)
    print(year)
    filtered_lake_df = lake_df[(lake_df['Month'] == month) & (lake_df['Year'] == year)]
    
    # Create the heatmap on the filtered data
    heatmap = create_heatmap(filtered_lake_df, 'heatmap.html')
    # Return statement as it is
    # Render the admin panel
    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    userType = session.get('userType')
    profile_image = session.get('profile_image')

    # Create the heatmap on the map
    #heatmap = create_heatmap(lake_df)
    #geojson_data1=geojson_data1, 
    return render_template('index.html',  heatmap=heatmap, lake_csv_data=lake_csv_data_with_wqi_json, user_id=user_id, fname=fname, lname=lname, userType=userType, profile_image=profile_image)

#charts
@app.route('/charts')
def charts():
    # Load data from the database
    db_data = WaterQuality.query.all()
    # Convert the database data to a list of dictionaries
    db_data_dict = []
    for record in db_data:
        # Convert the SQLAlchemy model object to a dictionary
        record_dict = {
            'id':record.id,
            'stationid': record.stationid,
            'Barangay': record.Barangay,
            'Month': record.Month,
            'Year': int(record.Year) if record.Year.strip() else None,  # Check for empty string
            'pH': float(record.pH) if record.pH.strip() else None,  # Convert to float, and check for empty string
            'Ammonia': float(record.Ammonia) if record.Ammonia.strip() else None,  # Convert to float, and check for empty string
            'DO': float(record.DO) if record.DO.strip() else None,  # Convert to float, and check for empty string
            'Nitrate': float(record.Nitrate) if record.Nitrate.strip() else None,  # Convert to float, and check for empty string
            'Phosphate': float(record.Phosphate) if record.Phosphate.strip() else None  # Convert to float, and check for empty string

        }

        db_data_dict.append(record_dict)

    # Convert the list of dictionaries to a Pandas DataFrame
    db_data_df = pd.DataFrame(db_data_dict)

    # Add more conversions as needed

    # Load data from the CSV file
    csv_data = pd.read_csv('taaldata.csv', encoding='ISO-8859-1')
    # Standardize the CSV data
    numerical_columns = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']
    for col in numerical_columns:
        csv_data[col] = csv_data[col].astype(float)

    # Calculate WQI for the CSV data and the database data
    db_data_df = calculate_wqi(db_data_df)
    csv_data = calculate_wqi(csv_data)

    for data in [db_data_df, csv_data]:
        for col in numerical_columns:
            data[col] = data[col].apply(lambda x: None if pd.isna(x) else x)

    # Combine the data from the database and CSV
    combined_data_df = pd.concat([db_data_df, csv_data], ignore_index=True)

    # Convert the combined data DataFrame to JSON format
    combined_data_json = combined_data_df.to_json(orient='records')

# Render the admin panel
    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    userType = session.get('userType')
    profile_image = session.get('profile_image')
    
    return render_template('charts.html', taal_json_data=combined_data_json, user_id=user_id, fname=fname, lname=lname, userType=userType, profile_image=profile_image)

@app.route('/logout')
def logout():
    if current_user.is_authenticated:
        logout_user()
        flash('Logged out successfully', 'info')
    else:
        flash('You are not logged in', 'warning')

    return redirect(url_for('index'))

@app.route('/save_to_database', methods=['POST'])
def save_to_database():
    if request.method == 'POST':
        print("Received a POST request to save data.")  # Add a debug message

        try:
            # Retrieve data from the form
            stationid = request.form.get('station_id')
            barangay = request.form.get('station_name')
            month = request.form.get('month')
            year = request.form.get('year')
            ph = request.form.get('ph')
            ammonia = request.form.get('ammonia')
            do = request.form.get('do')
            nitrate = request.form.get('nitrate')
            phosphate = request.form.get('phosphate')
            time = request.form.get('time')
            weather_condition = request.form.get('weater-condition')
            wind_direction = request.form.get('wind-direction')
            color_of_water = request.form.get('color-of-water')
            air_temperature = request.form.get('air-temperature')
            water_transparency = request.form.get('water-transparency')
            water_temperature = request.form.get('water-temp')

            # Create a new WaterQuality object and assign values
            water_quality = WaterQuality(
                stationid=stationid,
                Barangay=barangay,
                Month=month,
                Year=year,
                pH=ph,
                Ammonia=ammonia,
                DO=do,
                Nitrate=nitrate,
                Phosphate=phosphate,
                Time=time,
                WeatherCondition=weather_condition,
                WindDirection=wind_direction,
                ColorOfWater=color_of_water,
                AirTemperature=air_temperature,
                WaterTransparency=water_transparency,
                WaterTemperature=water_temperature
            )

            # Add the object to the session and commit the changes
            db.session.add(water_quality)
            db.session.commit()

            print("Data saved successfully.")  # Add a success message

            # Redirect to a different route (change 'route_name' to your desired route)
            return redirect(url_for('charts'))
        except Exception as e: # Log the error
            print("Error:", str(e))
            return jsonify({"message": "Error occurred while saving data", "status": "error"})
    else:
        print("Received a non-POST request to save_to_database.")  # Add a debug message for non-POST requests
        return jsonify({"message": "Invalid request method", "status": "error"})



#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-SAVE ANNOUNCEMENT=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
class Announcement(UserMixin, db.Model):
    __tablename__ = 'announcementtb'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    description = db.Column(db.String(255), nullable=False)
    status = db.Column(db.String(255), nullable=False)
    announcement_image = db.Column(db.String(255), nullable=True)

# Announcment
@app.route('/announcements')
def announcements():
      # Load data from the database
    db_data = Announcement.query.all()

    # Convert the database data to a list of dictionaries
    db_data_dict = [
        {
            'id': record.id,
            'title': record.title,
            'description': record.description,
            'status' :record.status,
            'announcement_image': record.announcement_image.replace('\\', '/')  # Replace backslashes with forward slashes
        }
        for record in db_data
    ]
     # Convert the list of dictionaries to a JSON string
    user_data_json = json.dumps(db_data_dict)
    #print(user_data_json)
    # user_id = session.get('id')
    # print(f'user_id: {user_id}')
    # fname = session.get('fname')
    # lname = session.get('lname')
    # status = session.get('status')
    # userType = session.get('userType')
    # profile_image = session.get('profile_image')
    #return render_template('usermanagement.html',user_data=user_data_json, id=user_id, fname=fname, lname=lname,status=status,userType=userType, profile_image=profile_image)
 
 # Render the admin panel
    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    userType = session.get('userType')
    profile_image = session.get('profile_image')
    
    return render_template('announcements.html', user_data=user_data_json, user_id=user_id, fname=fname, lname=lname, userType=userType, profile_image=profile_image)

@app.route('/update_announcement', methods=['POST'])
@login_required
def update_announcement():
    try:
        if request.method == 'POST':
            # Retrieve data from the form
            id = request.form.get('id')
            title = request.form.get('updatedTitle')
            description = request.form.get('updatedDescription')
            status = request.form.get('updatedStatus')

            # Find the announcement by ID
            announcement_to_update = Announcement.query.get(id)

            # Update announcement data if found
            if announcement_to_update:
                announcement_to_update.title = title
                announcement_to_update.description = description
                announcement_to_update.status = status

                # Handle file upload
                if 'updatedImage' in request.files:
                    updated_image = request.files['updatedImage']
                    if updated_image.filename != '':
                        # Process the uploaded image
                        if '.' in updated_image.filename and \
                                updated_image.filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS:
                            filename = secure_filename(updated_image.filename)
                            filepath = os.path.join(UPLOAD_FOLDER, filename)
                            updated_image.save(filepath)
                            # Update the image path in the database
                            announcement_to_update.announcement_image = filepath
                        else:
                            flash('Invalid file extension. Allowed extensions are png, jpg, jpeg, gif', 'error')
                            return redirect(request.url)
                    else:
                        # No image provided, set announcement_image to None
                        announcement_to_update.announcement_image = None

                # Commit the changes to the database
                db.session.commit()

                # Flash message for success
                flash('Announcement updated successfully', 'success')

                # Redirect to announcements page
                return redirect(url_for('announcements'))

            else:
                flash('Announcement not found', 'error')
                return redirect(url_for('announcements'))  # Redirect with an error flash message

    except Exception as e:
        flash('Error occurred while updating announcement', 'error')
        return redirect(url_for('announcements'))  # Redirect with an error flash message



UPLOAD_FOLDER = 'static/announcement'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

@app.route('/save_announcement', methods=['POST'])
def save_announcement():
    if request.method == 'POST':
        try:
            title = request.form.get('title')
            description = request.form.get('description')
            status = 'active'

            announce_data = Announcement(
                title=title,
                description=description,
                status=status
            )

            # Handle profile image upload
            if 'announce-image' in request.files:
                announcement_image = request.files['announce-image']
                if announcement_image.filename != '':
                    # Process the profile image as before
                    if '.' in announcement_image.filename and \
                            announcement_image.filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS:
                        filename = secure_filename(announcement_image.filename)
                        filepath = os.path.join(UPLOAD_FOLDER, filename)
                        announcement_image.save(filepath)
                        announce_data.announcement_image = filepath  # corrected assignment
                    else:
                        flash('Invalid file extension. Allowed extensions are png, jpg, jpeg, gif', 'error')
                        return redirect(request.url)
                else:
                    # No profile image provided, set announcement_image to None
                    announce_data.announcement_image = None
            else:
                # No profile image provided, set announcement_image to None
                announce_data.announcement_image = None


            db.session.add(announce_data)
            db.session.commit()

            flash('User created successfully', 'success')

            print("User data saved successfully")

            return redirect(url_for('announcements'))
        except Exception as e:
            print(f"Error occurred: {e}")
            db.session.rollback()
            flash('Error occurred while saving data', 'error')
            return redirect(request.url)

    else:
        return jsonify({"message": "Invalid request method", "status": "error"})

#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
    
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-SAVE SERVICES=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
class Services(db.Model):
    __tablename__ = 'servicestb'

    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), nullable=False)
    service_image = db.Column(db.String(255))  # Update this line


# Announcment
@app.route('/services')
def services():
    #   # Load data from the database
    db_data = Services.query.all()

    # # Convert the database data to a list of dictionaries
    db_data_dict = [
        {
            'id': record.id,
            'title': record.title,
            'description': record.description,
            'status' :record.status,
            'service_image': record.service_image.replace('\\', '/')  # Replace backslashes with forward slashes
        }
        for record in db_data
    ]
     # Convert the list of dictionaries to a JSON string
    user_data_json = json.dumps(db_data_dict)
    # Render the admin panel
    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    userType = session.get('userType')
    profile_image = session.get('profile_image')
    
    return render_template('services.html', user_data=user_data_json, user_id=user_id, fname=fname, lname=lname, userType=userType, profile_image=profile_image)

@app.route('/update_service', methods=['POST'])
@login_required
def update_service():
    try:
        # Retrieve data from the form
        id = request.form.get('id')
        title = request.form.get('updatedTitle')
        description = request.form.get('updatedDescription')
        status = request.form.get('updatedStatus')

        # Find the announcement by ID
        service_to_update = Services.query.get(id)

        # Check if the service exists
        if service_to_update:
            # Update service data
            service_to_update.title = title
            service_to_update.description = description
            service_to_update.status = status

            # Handle file upload
            if 'updatedImage' in request.files:
                updated_image = request.files['updatedImage']
                if updated_image.filename != '':
                    # Process the uploaded image
                    if '.' in updated_image.filename and \
                            updated_image.filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS:
                        filename = secure_filename(updated_image.filename)
                        filepath = os.path.join(UPLOAD_FOLDER, filename)
                        updated_image.save(filepath)
                        # Update the image path in the database
                        service_to_update.service_image = filepath
                    else:
                        flash('Invalid file extension. Allowed extensions are png, jpg, jpeg, gif', 'error')
                        return redirect(request.url)
                else:
                    # No image provided, set service_image to None
                    service_to_update.service_image = None

            # Commit the changes to the database
            db.session.commit()

            # Flash message for success
            flash('Service updated successfully', 'success')
        else:
            flash('Service not found', 'error')

    except Exception as e:
        flash(f'Error occurred while updating service: {e}', 'error')

    # Redirect to services page
        
        
    return redirect(url_for('services'))



UPLOAD_FOLDER = 'static/service'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

@app.route('/save_service', methods=['POST'])
def save_service():
    if request.method == 'POST':
        try:
            title = request.form.get('title')
            description = request.form.get('description')
            status = 'active'

            service_data = Services(  # Changed to Services instead of Announcement
                title=title,
                description=description,
                status=status
            )

            # Handle profile image upload
            if 'service-image' in request.files:
                service_image = request.files['service-image']
                if service_image.filename != '':
                    # Process the profile image as before
                    if '.' in service_image.filename and \
                            service_image.filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS:
                        filename = secure_filename(service_image.filename)
                        filepath = os.path.join(UPLOAD_FOLDER, filename)
                        service_image.save(filepath)
                        service_data.service_image = filepath  # Assign to service_image column
                    else:
                        flash('Invalid file extension. Allowed extensions are png, jpg, jpeg, gif', 'error')
                        return redirect(request.url)
                else:
                    # No profile image provided, set service_image to None
                    service_data.service_image = None
            else:
                # No profile image provided, set service_image to None
                service_data.service_image = None

            db.session.add(service_data)
            db.session.commit()

            flash('Service saved successfully', 'success')

            return redirect(url_for('services'))
        except Exception as e:
            print(f"Error occurred: {e}")
            db.session.rollback()
            flash('Error occurred while saving data', 'error')
            return redirect(request.url)

    else:
        return jsonify({"message": "Invalid request method", "status": "error"})


#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-

#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-SAVE ACTIVITIES=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
class Activity(db.Model):
    __tablename__ = 'activitytb'

    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), nullable=False)
    activity_image = db.Column(db.String(255))
    date = db.Column(db.Date, nullable=False)


# Announcment
@app.route('/activities')
def activities():
    #   # Load data from the database
    db_data = Activity.query.all()

    # # Convert the database data to a list of dictionaries
    db_data_dict = [
        {
            'id': record.id,
            'title': record.title,
            'description': record.description,
            'status' :record.status,
            'activity_image': record.activity_image.replace('\\', '/') if record.activity_image else '',  # Replace backslashes with forward slashes
            'date': record.date.isoformat() if record.date else ''   # Convert date to ISO 8601 format
        }
        for record in db_data
    ]
     # Convert the list of dictionaries to a JSON string
    user_data_json = json.dumps(db_data_dict)
    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    userType = session.get('userType')
    profile_image = session.get('profile_image')
    return render_template('activities.html', user_data=user_data_json, user_id=user_id, fname=fname, lname=lname, userType=userType, profile_image=profile_image)

@app.route('/update_activities', methods=['POST'])
@login_required
def update_activities():
    try:
        # Retrieve data from the form
        id = request.form.get('id')
        title = request.form.get('updatedTitle')
        description = request.form.get('updatedDescription')
        status = request.form.get('updatedStatus')
        date_str = request.form.get('updatedDate')

        # Find the announcement by ID
        activity_to_update = Activity.query.get(id)

        # Check if the service exists
        if activity_to_update:
            # Update service data
            activity_to_update.title = title
            activity_to_update.description = description
            activity_to_update.status = status
            activity_to_update.date=date_str

            # Handle file upload
            if 'updatedImage' in request.files:
                updated_image = request.files['updatedImage']
                if updated_image.filename != '':
                    # Process the uploaded image
                    if '.' in updated_image.filename and \
                            updated_image.filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS:
                        filename = secure_filename(updated_image.filename)
                        filepath = os.path.join(UPLOAD_FOLDER, filename)
                        updated_image.save(filepath)
                        # Update the image path in the database
                        activity_to_update.activity_image = filepath
                    else:
                        flash('Invalid file extension. Allowed extensions are png, jpg, jpeg, gif', 'error')
                        return redirect(request.url)
                else:
                    # No image provided, set service_image to None
                    activity_to_update.activity_image = None

            # Commit the changes to the database
            db.session.commit()

            # Flash message for success
            flash('Activity updated successfully', 'success')
        else:
            flash('Activity not found', 'error')

    except Exception as e:
        flash(f'Error occurred while updating activity: {e}', 'error')

    # Redirect to services page
    return redirect(url_for('activities'))



UPLOAD_FOLDER = 'static/activity'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

@app.route('/save_activities', methods=['POST'])
def save_activities():
    
    if request.method == 'POST':
        try:
           
            title = request.form.get('title')
            description = request.form.get('description')
            status = 'active'
            date_str = request.form.get('date')

            activity_data = Activity(  # Changed to Services instead of Announcement
                title=title,
                description=description,
                status=status,
                date=date_str
            )

            # Handle profile image upload
            if 'activity-image' in request.files:
                activity_image = request.files['activity-image']
                if activity_image.filename != '':
                    # Process the profile image as before
                    if '.' in activity_image.filename and \
                            activity_image.filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS:
                        filename = secure_filename(activity_image.filename)
                        filepath = os.path.join(UPLOAD_FOLDER, filename)
                        activity_image.save(filepath)
                        activity_data.activity_image = filepath  # Assign to service_image column
                    else:
                        flash('Invalid file extension. Allowed extensions are png, jpg, jpeg, gif', 'error')
                        return redirect(request.url)
                else:
                    # No profile image provided, set service_image to None
                    activity_data.activity_image = None
            else:
                # No profile image provided, set service_image to None
                activity_data.activity_image = None

            db.session.add(activity_data)
            db.session.commit()

            flash('Activity saved successfully', 'success')

            return redirect(url_for('activities'))
        except Exception as e:
            print(f"Error occurred: {e}")
            db.session.rollback()
            flash('Error occurred while saving data', 'error')
            return redirect(request.url)

    else:
        return jsonify({"message": "Invalid request method", "status": "error"})


#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
# Define the function to map months to quarters
@app.route('/update_data', methods=['POST'])
def update_data():
    try:
        with app.app_context():
            if request.method == 'POST':
                update_id = int(request.form['update_id1'])
                update_stationid=request.form['updatestation_id']
                update_station = request.form['updatestation_name']
                update_month = request.form['updateMonth']
                update_year = request.form['updateyear']
                update_do = float(request.form['updatedo'])
                update_ph = float(request.form['updateph'])
                update_phosphate = float(request.form['updatephosphate'])
                update_ammonia = float(request.form['updateammonia'])
                update_nitrate = float(request.form['updatenitrate'])

                # Retrieve the WaterQuality object by ID
                entry_to_update = WaterQuality.query.get(update_id)

                # Update object if found
                if entry_to_update:
                    entry_to_update.stationid = update_stationid
                    entry_to_update.Barangay = update_station
                    entry_to_update.Month = update_month
                    entry_to_update.Year = update_year
                    entry_to_update.DO = update_do
                    entry_to_update.pH = update_ph
                    entry_to_update.Phosphate = update_phosphate
                    entry_to_update.Ammonia = update_ammonia
                    entry_to_update.Nitrate = update_nitrate

                    # Commit the changes to the database
                    db.session.commit()

                    flash('Data updated successfully', 'success')
                    return redirect(url_for('charts'))

                else:
                    flash('Data not found', 'error')
                    return redirect(url_for('charts', error_message='Data not found'))

    except Exception as e:
        flash('Error occurred while updating data', 'error')
        return redirect(url_for('charts', error_message='Error occurred while updating data'))
#
#
#
#
def map_month_to_quarter(month):
    quarter_map = {
        1: 'Q1', 2: 'Q1', 3: 'Q1',
        4: 'Q2', 5: 'Q2', 6: 'Q2',
        7: 'Q3', 8: 'Q3', 9: 'Q3',
        10: 'Q4', 11: 'Q4', 12: 'Q4'
    }
    return quarter_map.get(month, 'Unknown')

# Calculate WQI and add 'Quarter' column to DataFrame
def calculate_and_map_quarters(df):
    df = calculate_wqi(df)
    df['Quarter'] = df['Month'].apply(map_month_to_quarter)
    return df

@app.route('/load_map')
def load_map():
   # Read shapefile data using GeoPandas for the first shapefile only
    shapefile_path1 = r'shapefile/taalLake/Taal-Lake.shp'
    data1 = gpd.read_file(shapefile_path1)
    geojson_data1 = data1.to_json()

    lake_df = pd.read_csv('templates/taaldata.csv', encoding='ISO-8859-1')
    lake_df = calculate_wqi(lake_df)  # Calculate WQI
    lake_csv_data_with_wqi_json = lake_df.to_json(orient='records')

    # Return statement as it is
    return render_template('gis.html', lake_csv_data=lake_csv_data_with_wqi_json, geojson_data1=geojson_data1)

#---------------------------------------- reading csv data
def read_csv_data(filename):
    try:
        data = pd.read_csv(filename)
        data = calculate_wqi(data)  # Calculate WQI and WQC values
        return data.to_json()
    except Exception as e:
        print('Error reading data from CSV:', e)
        return None

#----------------------------------------------------INDEX----------------------------------------@app.route('/')

@app.route('/')
def index():
    # Load data from the database
    db_data = WaterQuality.query.all()
    # Convert the database data to a list of dictionaries
    db_data_dict = []
    for record in db_data:
        # Convert the SQLAlchemy model object to a dictionary
        record_dict = {
            'stationid': record.stationid,
            'Barangay': record.Barangay,
            'Month': record.Month,
            # 'Year': int(record.Year) if record.Year is not None else None,  # Check for None
            # 'pH': float(record.pH) if record.pH is not None else None,  # Convert to float, and check for None
            # 'Ammonia': float(record.Ammonia) if record.Ammonia is not None else None,  # Convert to float, and check for None
            # 'DO': float(record.DO) if record.DO is not None else None,  # Convert to float, and check for None
            # 'Nitrate': float(record.Nitrate) if record.Nitrate is not None else None,  # Convert to float, and check for None
            # 'Phosphate': float(record.Phosphate) if record.Phosphate is not None else None  # Convert to float, and check for None
            'Year': int(record.Year) if record.Year.strip() else None,  # Check for empty string
            'pH': float(record.pH) if record.pH.strip() else None,  # Convert to float, and check for empty string
            'Ammonia': float(record.Ammonia) if record.Ammonia.strip() else None,  # Convert to float, and check for empty string
            'DO': float(record.DO) if record.DO.strip() else None,  # Convert to float, and check for empty string
            'Nitrate': float(record.Nitrate) if record.Nitrate.strip() else None,  # Convert to float, and check for empty string
            'Phosphate': float(record.Phosphate) if record.Phosphate.strip() else None  # Convert to float, and check for empty string
        }

        db_data_dict.append(record_dict)

    # Convert the list of dictionaries to a Pandas DataFrame
    db_data_df = pd.DataFrame(db_data_dict)

    # Add more conversions as needed

    # Load data from the CSV file
    csv_data = pd.read_csv('taaldata.csv', encoding='ISO-8859-1')
    # Standardize the CSV data
    numerical_columns = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']
    for col in numerical_columns:
        csv_data[col] = csv_data[col].astype(float)

    # Calculate WQI for the CSV data and the database data
    db_data_df = calculate_wqi(db_data_df)
    csv_data = calculate_wqi(csv_data)

    for data in [db_data_df, csv_data]:
        for col in numerical_columns:
            data[col] = data[col].apply(lambda x: None if pd.isna(x) else x)

    # Combine the data from the database and CSV
    combined_data_df = pd.concat([db_data_df, csv_data], ignore_index=True)

    # Convert the combined data DataFrame to JSON format
    combined_data_json = combined_data_df.to_json(orient='records')
 
# Render the admin panel
    shapefile_path1 = r'shapefile/taalLake/Taal-Lake.shp'
    data1 = gpd.read_file(shapefile_path1)
    geojson_data1 = data1.to_json()

    lake_df = pd.read_csv('taaldata.csv', encoding='ISO-8859-1')
    lake_df = lake_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    lake_df = calculate_wqi(lake_df)  # Calculate WQI
    lake_csv_data_with_wqi_json = lake_df.to_json(orient='records')
    
    # Return statement as it is
   # Filter the data for January 2018
    filtered_data = combined_data_df[(combined_data_df['Month'] == 'January ') & (combined_data_df['Year'] == 2018)]
    print('dao chart')
    print(filtered_data)
    # Get selected month and year from the request query parameters
    selected_month = 'January'
    selected_year = 2018

    # Filter data based on selected month and year
    # filtered_data = combined_data_df[(combined_data_df['Month'] == selected_month) & (combined_data_df['Year'] == int(selected_year))]#

    # Extract barangays and parameter values
    barangays = filtered_data['Barangay'].unique()
    parameters = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']

    parameter_values = {param: [] for param in parameters}
    parameter_suitability = {param: [] for param in parameters}

    for barangay in barangays:
        barangay_data = filtered_data[filtered_data['Barangay'] == barangay]
        for param in parameters:
            values = barangay_data[param].values
            suitability = [check_suitability(value, param)['suitability'] for value in values]
            parameter_values[param].append(values.mean())
            parameter_suitability[param].append(suitability[0])

    # Plot the data
    plt.figure(figsize=(12, 6))

    bar_width = 0.15
    index = np.arange(len(barangays))

    # Plot bars for each parameter
    for i, param in enumerate(parameters):
        offset = bar_width * (i - len(parameters) // 2)
        colors = ['blue' if s == 'Suitable' else 'red' for s in parameter_suitability[param]]
        bars = plt.bar(index + offset, parameter_values[param], bar_width, label=param, color=colors)
        for bar in bars:
            height = bar.get_height()
            plt.annotate(f'{param}', xy=(bar.get_x() + bar.get_width() / 2, height),
                         xytext=(0, 3),  # 3 points vertical offset
                         textcoords="offset points",
                         ha='center', va='bottom', rotation=90)  # Rotate label by 90 degrees
            
    green_patch = mpatches.Patch(color='blue', label='Normal')
    red_patch = mpatches.Patch(color='red', label='Warning')
    plt.legend(handles=[green_patch, red_patch])
    plt.xlabel('Barangay')
    plt.ylabel('Value')
    plt.title(f'Water Quality Parameters for {selected_month} {selected_year}')
    plt.xticks(index, barangays, rotation=45, ha='right')
    plt.tight_layout()

    # Save the plot as an image
    plot_path = f'static/bar_chart_1_{selected_month}_{selected_year}.png'
    plt.savefig(plot_path)
    plt.close()


    # Close the plot to free up resources
    plt.close()

#############################################################
        # Create a DataFrame from the sample data
    combined_data_df = pd.read_csv('taaldata.csv', encoding='ISO-8859-1')
    combined_data_df = calculate_wqi(combined_data_df)
    # Filter the data for January 2018
    years = range(2018, 2025)
    wqcs = ['Excellent', 'Good', 'Poor', 'Very Poor', 'Unsuitable']  # All possible categories

    # Define colors for WQI classifications
    colors = {'Unsuitable': 'blue', 'Very Poor': 'orange', 'Poor': 'green', 'Good': 'red', 'Excellent': 'purple'}

    # Calculate percentage of each WQI classification for each year
    wqc_percentages = {}
    for year in years:
        year_data = combined_data_df[combined_data_df['Year'] == year]
        wqc_counts = year_data['wqc'].value_counts()
        total = wqc_counts.sum()
        wqc_percentages[year] = {wqc: (count / total) * 100 for wqc, count in wqc_counts.items()}

    # Plot the data
    plt.figure(figsize=(12, 6))

    bar_width = 0.15
    index = np.arange(len(years))

    # Plot bars for each WQI classification
    for i, wqc in enumerate(wqcs):
        offset = bar_width * (i - len(wqcs) // 2)
        percentages = [wqc_percentages[year].get(wqc, 0) for year in years]
        plt.bar(index + offset, percentages, bar_width, label=wqc, color=colors.get(wqc, 'gray'))

    # Set labels, title, and legend with corresponding colors and descriptions
    plt.xlabel('Year')
    plt.ylabel('Percentage')
    plt.title('Water Quality Index Classification by Year')
    plt.xticks(index, years)

    # Generate legend handles and labels dynamically based on all possible WQI classifications
    legend_handles = [plt.Rectangle((0,0),1,1, color=colors[wqc]) for wqc in wqcs if wqc in colors]
    plt.legend(legend_handles, wqcs, title='WQI Classification')

    # Save the plot as an image
    plot_path1 = 'static/year_wqc_percentage_plot.png'
    plt.savefig(plot_path1)
    #plt.close()
    # Convert the combined data DataFrame to JSON format
    combined_data_json = combined_data_df.to_json(orient='records')

    activities = Activity.query.limit(6).all()
    # Retrieve contact data from the database
    contacts = Contact.query.all()
    services = Services.query.order_by(desc(Services.id)).limit(9).all()
    # Render the index.html with the bar chart image path
    announcements = Announcement.query.order_by(desc(Announcement.id)).limit(2).all()
    print('this it the csv')
    print(lake_df)
    # Get month and year parameters from the request
    month = request.args.get('month')
    year = request.args.get('year')

    # Check if month and year are not None
    if month is not None and year is not None:
        # Convert month and year to integers
        month = month
        year = int(year)
    else:
        # Set default values to January 2019
        month = 'January'
        year = 2018

    # Filter lake_df based on month and year
    filtered_lake_df = lake_df[(lake_df['Month'] == month) & (lake_df['Year'] == year)]
    print(filtered_lake_df)
    # Create the heatmap on the filtered data
    heatmap = create_heatmap(filtered_lake_df, 'heatmap.html')
    return render_template('index.html', heatmap=heatmap, taal_json_data=combined_data_json, lake_csv_data=lake_csv_data_with_wqi_json, geojson_data1=geojson_data1, plot_path=plot_path,plot_path1=plot_path1,activities=activities,contacts=contacts, services=services, announcements=announcements)


#=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-



@app.route('/send_email', methods=['POST'])
def send_email():
    recipient_email = request.form.get('email')
    full_name = request.form.get('full_name')
    subject = request.form.get('subject')
    message_body = request.form.get('message')
    print(f"Recipient Email: {recipient_email}")
    print(f"Subject: {subject}")
    print(f"Message Body: {message_body}")
    if recipient_email and full_name and subject and message_body:
        sender_email = 'luigitimola@gmail.com'  # Replace with your email
        app_password = 'yvmw udfs zijy iprs'  # Replace with your app password

        # Create a message object
        message = MIMEMultipart()
        message['From'] = sender_email
        message['To'] = 'marcluigitimola@gmail.com'
        message['Subject'] = subject
        body = f'Full Name: {full_name}\n\nMessage: {message_body}\n\nRecipient Email: {recipient_email}'
        message.attach(MIMEText(body, 'plain'))

        try:
            # Establish an SMTP connection
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.set_debuglevel(1)  # Enable SMTP debug
                # Log in to the SMTP server
                server.login(sender_email, 'yvmw udfs zijy iprs')
                # Send the email
                server.sendmail(sender_email, recipient_email, message.as_string())

            flash('Email sent successfully!', 'success')
        except Exception as e:
            flash(f'Error sending email: {str(e)}', 'error')
    else:
        flash('Please fill in all the fields.', 'error')

    return redirect(url_for('index'))

#
#
#
@app.route('/forgot_password', methods=['POST'])
def forgot_password():
    recipient_email = request.form.get('email')

    # Check if the email exists in the database
    user = User.query.filter_by(email=recipient_email).first()

    if user:
        # Email exists, proceed with sending the reset email
        sender_email = 'hyperaktib@gmail.com'  # Replace with your email
        subject = 'Password Reset'
        temporary_password = generate_temporary_password()

        # Hash the temporary password before saving it in the database
        hashed_temporary_password = bcrypt.generate_password_hash(temporary_password).decode('utf-8')

        # Update the user's password with the hashed temporary password
        user.password = hashed_temporary_password
        db.session.commit()

        message = MIMEMultipart()
        message['From'] = sender_email
        message['To'] = recipient_email
        message['Subject'] = subject
        body = f'Your temporary password: {temporary_password}'
        message.attach(MIMEText(body, 'plain'))

        try:
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login(sender_email, 'nkuc lovx yegy ssdm')  # Replace with your Gmail app password
                server.sendmail(sender_email, recipient_email, message.as_string())

            flash('Password reset email sent successfully!', 'success')
        except Exception as e:
            flash(f'Error sending email: {str(e)}', 'error')
    else:
        flash('Email not found in the database.', 'error')

    return redirect(url_for('login'))


def generate_temporary_password():
    # Implement your logic to generate a random temporary password
    # For example, you can use the secrets module to generate a secure token
    import secrets
    return secrets.token_urlsafe(12)
#=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
#=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=--=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-==-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-
def is_float(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

@app.route('/update_chart_data', methods=['GET'])
def update_chart_data():
    # Load data from the database
    db_data = WaterQuality.query.all()

    # Convert the database data to a list of dictionaries
    db_data_dict = []
    for record in db_data:
        # Convert the SQLAlchemy model object to a dictionary
        record_dict = {
            'Month': record.Month,
            'Year': int(record.Year) if record.Year and record.Year.strip().isdigit() else None,
            'pH': float(record.pH) if record.pH and record.pH.strip() and is_float(record.pH) else None,
            'Ammonia': float(record.Ammonia) if record.Ammonia and record.Ammonia.strip() and is_float(record.Ammonia) else None,
            'DO': float(record.DO) if record.DO and record.DO.strip() and is_float(record.DO) else None,
            'Nitrate': float(record.Nitrate) if record.Nitrate and record.Nitrate.strip() and is_float(record.Nitrate) else None,
            'Phosphate': float(record.Phosphate) if record.Phosphate and record.Phosphate.strip() and is_float(record.Phosphate) else None,
        }
        db_data_dict.append(record_dict)

    # Convert the list of dictionaries to a Pandas DataFrame
    db_data_df = pd.DataFrame(db_data_dict)

    # Load data from the CSV file
    csv_data = pd.read_csv('taaldata.csv', encoding='ISO-8859-1')
    csv_data = csv_data.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    # Standardize the CSV data
    numerical_columns = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']
    for col in numerical_columns:
        csv_data[col] = csv_data[col].astype(float)

    # Calculate WQI for the CSV data and the database data
    db_data_df = calculate_wqi(db_data_df)
    csv_data = calculate_wqi(csv_data)

    for data in [db_data_df, csv_data]:
        for col in numerical_columns:
            data[col] = data[col].apply(lambda x: None if pd.isna(x) else x)

    # Combine the data from the database and CSV
    combined_data_df = pd.concat([db_data_df, csv_data], ignore_index=True)

    # Get selected year from the request query parameters
    selected_year = request.args.get('year')
    # Filter data based on selected year
    if selected_year.lower() == 'all':
        years = range(2018, 2025)
        wqcs = ['Excellent', 'Good', 'Poor', 'Very Poor', 'Unsuitable']  # All possible categories

        # Define colors for WQI classifications
        colors = {'Unsuitable': 'blue', 'Very Poor': 'orange', 'Poor': 'green', 'Good': 'red', 'Excellent': 'purple'}

        # Calculate percentage of each WQI classification for each year
        wqc_percentages = {}
        for year in years:
            year_data = combined_data_df[combined_data_df['Year'] == year]
            wqc_counts = year_data['wqc'].value_counts()
            total = wqc_counts.sum()
            wqc_percentages[year] = {wqc: (count / total) * 100 for wqc, count in wqc_counts.items()}

        # Plot the data
        plt.figure(figsize=(12, 6))

        bar_width = 0.15
        index = np.arange(len(years))

        # Plot bars for each WQI classification
        for i, wqc in enumerate(wqcs):
            offset = bar_width * (i - len(wqcs) // 2)
            percentages = [wqc_percentages[year].get(wqc, 0) for year in years]
            plt.bar(index + offset, percentages, bar_width, label=wqc, color=colors.get(wqc, 'gray'))

        # Set labels, title, and legend with corresponding colors and descriptions
        plt.xlabel('Year')
        plt.ylabel('Percentage')
        plt.title('Water Quality Index Classification by Year')
        plt.xticks(index, years)

        # Generate legend handles and labels dynamically based on all possible WQI classifications
        legend_handles = [plt.Rectangle((0,0),1,1, color=colors[wqc]) for wqc in wqcs if wqc in colors]
        plt.legend(legend_handles, wqcs, title='WQI Classification')

        # Save the plot as an image
        plot_path1 = 'static/year_wqc_percentage_plot1.png'
        plt.savefig(plot_path1)
    else:
        selected_year = int(request.args.get('year'))  # Convert selected_year to integer
        # Filter data based on selected year
        filtered_data = combined_data_df[combined_data_df['Year'] == selected_year]

        # Drop rows with NaN values in the 'Barangay' column
        filtered_data = filtered_data.dropna(subset=['Barangay'])

        # Get unique barangays for the selected year
        barangays = filtered_data['Barangay'].unique()

        # Define WQI categories
        wqcs = ['Excellent', 'Good', 'Poor', 'Very Poor', 'Unsuitable']

        # Define colors for WQI classifications
        colors = {'Unsuitable': 'blue', 'Very Poor': 'orange', 'Poor': 'green', 'Good': 'red', 'Excellent': 'purple'}

        # Plot the data
        plt.figure(figsize=(12, 6))

        for i, barangay in enumerate(barangays):
            # Filter data for the current barangay
            barangay_data = filtered_data[filtered_data['Barangay'] == barangay]

            # Calculate percentage of each WQI classification for the current barangay
            wqc_counts = barangay_data['wqc'].value_counts()
            total = wqc_counts.sum()
            wqc_percentages = {wqc: (count / total) * 100 for wqc, count in wqc_counts.items()}

            # Plot bars for each WQI classification for the current barangay
            bar_colors = [colors.get(wqc, 'gray') for wqc in wqcs]
            plt.bar([i + (j - len(wqcs) / 2) * 0.15 for j in range(len(wqcs))], 
                    [wqc_percentages.get(wqc, 0) for wqc in wqcs], 
                    width=0.15, label=barangay, color=bar_colors)

        # Set labels, title, and tick labels
        plt.xlabel('Water Quality Index Classification')
        plt.ylabel('Percentage')
        plt.title('Water Quality Index Classification for the Year {}'.format(selected_year))
        plt.xticks(range(len(wqcs)), wqcs)  # Use wqcs for x-axis ticks

        # Generate legend handles and labels dynamically based on all WQI classifications
        legend_handles = [plt.Rectangle((0, 0), 1, 1, color=colors.get(wqc, 'gray')) for wqc in wqcs]
        plt.legend(legend_handles, wqcs, title='WQI Classification', loc='upper right')

        # Set the tick labels for barangays
        plt.gca().set_xticks(range(len(barangays)))
        plt.gca().set_xticklabels(barangays, rotation=45, ha='right')

        # Save the plot as an image
        plot_path1 = f'static/year_wqc_percentage_plot_{selected_year}.png'
        plt.savefig(plot_path1)

        # Close the plot to free up resources
        plt.close()

    # Return the path to the plot image
    return jsonify({"plot_path": plot_path1})

@app.route('/update_dao_chart_data', methods=['GET'])
def update_dao_chart_data():
    # Load data from the database
    db_data = WaterQuality.query.all()

    # Convert the database data to a list of dictionaries
    db_data_dict = []
    for record in db_data:
        # Convert the SQLAlchemy model object to a dictionary
        record_dict = {
            'Month': record.Month,
            'Year': int(record.Year) if record.Year.strip() else None,
            'pH': float(record.pH) if record.pH.strip() else None,
            'Ammonia': float(record.Ammonia) if record.Ammonia.strip() else None,
            'DO': float(record.DO) if record.DO.strip() else None,
            'Nitrate': float(record.Nitrate) if record.Nitrate.strip() else None,
            'Phosphate': float(record.Phosphate) if record.Phosphate.strip() else None
        }
        db_data_dict.append(record_dict)

    # Convert the list of dictionaries to a Pandas DataFrame
    db_data_df = pd.DataFrame(db_data_dict)

    # Load data from the CSV file
    csv_data = pd.read_csv('taaldata.csv', encoding='ISO-8859-1')
    # Standardize the CSV data
    numerical_columns = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']
    for col in numerical_columns:
        csv_data[col] = csv_data[col].astype(float)

    # Calculate WQI for the CSV data and the database data
    db_data_df = calculate_wqi(db_data_df)
    csv_data = calculate_wqi(csv_data)

    for data in [db_data_df, csv_data]:
        for col in numerical_columns:
            data[col] = data[col].apply(lambda x: None if pd.isna(x) else x)

    # Combine the data from the database and CSV
    combined_data_df = pd.concat([db_data_df, csv_data], ignore_index=True)

    # Get selected month and year from the request query parameters
    selected_month = request.args.get('month')
    selected_year = request.args.get('year')

    # Filter data based on selected month and year
    filtered_data = combined_data_df[(combined_data_df['Month'] == selected_month) & (combined_data_df['Year'] == int(selected_year))]

    # Extract barangays and parameter values
    barangays = filtered_data['Barangay'].unique()
    parameters = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']

    parameter_values = {param: [] for param in parameters}
    parameter_suitability = {param: [] for param in parameters}

    for barangay in barangays:
        barangay_data = filtered_data[filtered_data['Barangay'] == barangay]
        for param in parameters:
            values = barangay_data[param].values
            suitability = [check_suitability(value, param)['suitability'] for value in values]
            parameter_values[param].append(values.mean())
            parameter_suitability[param].append(suitability[0])

    # Plot the data
    plt.figure(figsize=(12, 6))

    bar_width = 0.15
    index = np.arange(len(barangays))

    # Plot bars for each parameter
    for i, param in enumerate(parameters):
        offset = bar_width * (i - len(parameters) // 2)
        colors = ['blue' if s == 'Suitable' else 'red' for s in parameter_suitability[param]]
        bars = plt.bar(index + offset, parameter_values[param], bar_width, label=param, color=colors)
        for bar in bars:
            height = bar.get_height()
            plt.annotate(f'{param}', xy=(bar.get_x() + bar.get_width() / 2, height),
                         xytext=(0, 3),  # 3 points vertical offset
                         textcoords="offset points",
                         ha='center', va='bottom', rotation=90)  # Rotate label by 90 degrees

    green_patch = mpatches.Patch(color='blue', label='Normal')
    red_patch = mpatches.Patch(color='red', label='Warning')
    plt.legend(handles=[green_patch, red_patch])
    plt.xlabel('Barangay') 
    plt.ylabel('Value')
    plt.title(f'Water Quality Parameters for {selected_month} {selected_year}')
    plt.xticks(index, barangays, rotation=45, ha='right')
    plt.tight_layout()

    # Save the plot as an image
    plot_path1 = f'static/bar_chart_1_{selected_month}_{selected_year}.png'
    plt.savefig(plot_path1)
    plt.close()

    # Return the path to the plot image
    return jsonify({"plot_path": plot_path1})




@app.route('/index1')
def index1():
    # Load data from the database
    db_data = WaterQuality.query.all()
    # Convert the database data to a list of dictionaries
    db_data_dict = []
    for record in db_data:
        # Convert the SQLAlchemy model object to a dictionary
        record_dict = {
            'stationid': record.stationid,
            'Barangay': record.Barangay,
            'Month': record.Month,
            'Year': int(record.Year) if record.Year is not None else None,  # Check for None
            'pH': float(record.pH) if record.pH is not None else None,  # Convert to float, and check for None
            'Ammonia': float(record.Ammonia) if record.Ammonia is not None else None,  # Convert to float, and check for None
            'DO': float(record.DO) if record.DO is not None else None,  # Convert to float, and check for None
            'Nitrate': float(record.Nitrate) if record.Nitrate is not None else None,  # Convert to float, and check for None
            'Phosphate': float(record.Phosphate) if record.Phosphate is not None else None  # Convert to float, and check for None
        }

        db_data_dict.append(record_dict)

# After loading data from the database
    print("Database Data:")
    for record_dict in db_data_dict:
        print(record_dict)
    # Convert the list of dictionaries to a Pandas DataFrame
    db_data_df = pd.DataFrame(db_data_dict)

    # Add more conversions as needed

    # Load data from the CSV file
    csv_data = pd.read_csv('templates/taaldata.csv', encoding='ISO-8859-1')
    # Standardize the CSV data
    numerical_columns = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']
    for col in numerical_columns:
        csv_data[col] = csv_data[col].astype(float)

    # Calculate WQI for the CSV data and the database data
    db_data_df = calculate_wqi(db_data_df)
    csv_data = calculate_wqi(csv_data)

    for data in [db_data_df, csv_data]:
        for col in numerical_columns:
            data[col] = data[col].apply(lambda x: None if pd.isna(x) else x)

    # Combine the data from the database and CSV
    combined_data_df = pd.concat([db_data_df, csv_data], ignore_index=True)

    # Convert the combined data DataFrame to JSON format
    combined_data_json = combined_data_df.to_json(orient='records')

    return render_template('index1.html', taal_json_data=combined_data_json)


# ...

#-------------------------------------------------------reading data from csv-------------------------------------
@app.route('/taaldata.csv')#----------------------------CHANGE CSV HERE-------------------------------------
def serve_lake_csv():
    try:
        return send_from_directory(os.getcwd(), 'templates/taaldata.csv', as_attachment=True)
    except Exception as e:
        print('Error serving taaldata.csv:', e)
        return 'Error serving taaldata.csv', 404

    
#------------------------------------------------------WQI and WQC Computation -------------------------------------------------------------------
def calculate_wqi(df):
     # Assigned standard values based on WHO
    for col in [ 'DO', 'pH', 'Phosphate', 'Ammonia', 'Nitrate']:
        df[col] = df[col].fillna(round(df[col].mean(), 2))
    # Assigned standard values based on WHO
    sv_do = 5
    sv_pH = 8
    sv_phosphates = 0.5
    sv_ammonia = 0.5
    sv_nitrate = 7

    #K = 1 / ((1 / sv_do) + (1 / sv_pH) + (1 / sv_phosphates) + (1 / sv_ammonia) + (1 / sv_nitrate))

    # df['WiDO'] = K / sv_do
    # df['WipH'] = K / sv_pH
    # df['WiPhosphate'] = K / sv_phosphates
    # df['WiAmmonia'] = K / sv_ammonia
    # df['WiNitrate'] = K / sv_nitrate

    df['WiDO'] = 1 / sv_do
    df['WipH'] = 1 / sv_pH
    df['WiPhosphate'] = 1 / sv_phosphates
    df['WiAmmonia'] = 1 / sv_ammonia
    df['WiNitrate'] = 1 / sv_nitrate

    # Calculate individual quality indices (qi) for each parameter
    # qi_do = 100 * ((df['DO'] - 14.6) / (sv_do - 14.6))
    # qi_ph = 100 * ((df['pH'] - 7) / (sv_pH - 7))
    # qi_phosphates = 100 * ((df['Phosphate'] - 0) / (sv_phosphates - 0))
    # qi_ammonia = 100 * ((df['Ammonia'] - 0) / (sv_ammonia - 0))
    # qi_nitrate = 100 * ((df['Nitrate'] - 0) / (sv_nitrate - 0))

    qi_do = 100 * (df['DO']/sv_do)
    qi_ph = 100 * (df['pH']/sv_pH)
    qi_phosphates = 100 * (df['Phosphate']/sv_phosphates)
    qi_ammonia = 100 * (df['Ammonia']/sv_ammonia)
    qi_nitrate = 100 * (df['Nitrate']/sv_nitrate)


    # Calculate weighted quality indices (qi x wi) for each parameter
    xdo = qi_do * df.WiDO
    xph = qi_ph * df.WipH
    xphosphates = qi_phosphates * df.WiPhosphate
    xammonia = qi_ammonia * df.WiAmmonia
    xnitrate = qi_nitrate * df.WiNitrate

    # Calculate sum of (qi x wi)
    sum_qiwi = xdo + xph + xphosphates + xammonia + xnitrate

    # Calculate sum of wi
    sum_wi = df.WiDO + df.WipH + df.WiPhosphate + df.WiAmmonia + df.WiNitrate

    # Calculate Water Quality Index (wqi)
    df['wqi'] = round(sum_qiwi / sum_wi, 2)

    # Determine Water Quality Class (wqc)
    df['wqc'] = df['wqi'].apply(lambda x: ('Excellent' if (50 >= x >= 0)
                                          else ('Good' if (100 >= x >= 51)
                                                else ('Poor' if (200 >= x >= 101)
                                                      else ('Very Poor' if (300 >= x >= 201)
                                                            else 'Unsuitable')))))

    #print("DataFrame after WQI calculation:")
    #print(df)
    return df


#-------------------------------------------------- FPDF GENERATE REPORT ---------------------------------------------

# Assuming you have a 'Month' column in your river_df
# Map months to quarters and add a 'Quarter' column

# ... (other routes and functions)

from datetime import datetime

def create_header(pdf, selected_quarter, selected_year):
    quarter_to_months = {
            'Q1': 'January to March',
            'Q2': 'April to June',
            'Q3': 'July to September',
            'Q4': 'October to December'
        }
    quarter_range = quarter_to_months.get(selected_quarter, 'Unknown Quarter')
    # header_text = f"{quarter_range} {selected_year}"
    
    # Get the current date
    current_date = datetime.now()

    # Format the date
    formatted_date = current_date.strftime("%B %d, %Y")
    
    pdf.image('./static/img/bagong_pilipinas.png', 10, 10, 23, 20)
    pdf.image('./static/img/dabfar.png', 33, 7, 30, 30)
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 3, 'Republic of the Philippines', 0, 1, 'C')
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 4, 'Department of Agriculture', 0, 1, 'C')
    pdf.set_font('Arial', 'B', 8)
    pdf.cell(0, 4, 'BUREAU OF FISHERIES AND AQUATIC RESOURCES', 0, 1, 'C')
    pdf.cell(0, 4, 'REGIONAL OFFICE NO. 4-A', 0, 1, 'C')
    pdf.cell(0, 4, 'BATANGAS INLAND FISHERIES TECHNOLOGY', 0, 1, 'C')
    pdf.cell(0, 4, 'OUTREACH STATION', 0, 1, 'C')
    pdf.set_font('Arial', '', 8)
    pdf.set_text_color(0, 0, 255)
    pdf.cell(0, 3, 'bfar4abiftos@gmail.com.ph', 0, 1, 'C')
    pdf.set_text_color(0, 0, 0)
    pdf.ln(8)  # Add some space after the header
    
    pdf.set_font('Times', 'B', 10)
    pdf.cell(0, 4, 'Water Quality Updates in Taal Lake as of ' + selected_quarter +' '+str(selected_year), 0, 1, 'C')
    
    pdf.ln(3)
    
# ... (other imports and code)
def get_cell_color(parameter, value):
    if parameter == "pH":
        if 6.5 <= value <= 8.5:
            return (0, 0, 255)  # Blue
        elif 6.5 <= value <= 9.0:
            return (255, 255, 0)  # Yellow
        elif 6.0 <= value <= 9.0:
            return (255, 0, 0)  # Red
        elif 0 <= value < 6 or value > 9:
            return (0, 0, 0)  # Black
    # Define similar color logic for other parameters here
    return (0, 0, 0)  # Default color

# ... (other imports and code)

# Define a function to convert hex color codes to RGB values
def hex_to_rgb(hex_color):
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
# ... (other imports and code)

@app.route('/generate_excel')
def generate_excel():
    # Define sample data
    monitoring_stations = [
        {"Monitoring Station": "Station 1", "DO": 5.0, "pH": 7.0, "Phosphate": 0.02, "Ammonia": 0.05, "Nitrate": 10},
        {"Monitoring Station": "Station 2", "DO": 4.0, "pH": 6.5, "Phosphate": 0.03, "Ammonia": 0.08, "Nitrate": 15},
        # Add more stations and their data as needed
    ]

    # Create Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write header row
    header = ["Monitoring Station", "DO", "pH", "Phosphate", "Ammonia", "Nitrate"]
    worksheet.append(header)

    # Write data rows
    for station in monitoring_stations:
        row_data = [station[col] for col in header]
        worksheet.append(row_data)

    # Save workbook to file
    file_path = "river_report.xlsx"
    workbook.save(file_path)

    # Return the Excel file as a response
    return send_file(file_path, as_attachment=True, attachment_filename="river_report.xlsx")


month_to_number = {
    'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
    'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12
}

station_barangays = {
    'Talisay': ['Sampaloc', 'Quiling', 'San Isidro'],
    'Laurel': ['Buso-Buso', 'Leviste'],
    'Agoncillo': ['Bañaga', 'Manalaw'],
    'Mataas na Kahoy': ['Mataas na Kahoy', 'Nangkaan'],
    'Tanauan': ['Tanauan']
}

@app.route('/generate_report', methods=['POST'])
def generate_report():
    selected_month_number = request.form.get('month')  # Get selected month from form
    selected_year = int(request.form.get('year'))  # Get selected year from form
    #selected_month_number = month_to_number.get(selected_month, None)
    selected_stations = request.form.getlist('stations')
    brgy_count = 0
    station_counts = {
        'Talisay': 3,
        'Laurel': 2,
        'Agoncillo': 2,
        'Mataas na Kahoy': 2,
        'Tanauan': 1
    }
    
    # Initialize the variable to store selected barangays
    selected_barangays_list = []
    
    # Count the number of barangays per selected station
    for station in selected_stations:
        barangays = station_barangays.get(station, [])
        selected_barangays_list.extend(barangays)
    
    print(selected_barangays_list)
    

    # Iterate through selected stations and update brgy_count
    for station in selected_stations:
        if station in station_counts:
            brgy_count += station_counts[station]

    # Print or use brgy_count as needed
    # print("Total brgy_count:", brgy_count)
    
    # if selected_month_number is None:
    #     return "Invalid month selected"

    # if selected_year <= 2022:
    db_data = WaterQuality.query.all()
    # Convert the database data to a list of dictionaries
    db_data_dict = []
    for record in db_data:
        # Convert the SQLAlchemy model object to a dictionary
        record_dict = {
            'stationid': record.stationid,
            'Barangay': record.Barangay,
            'Month': record.Month,
            # 'Year': int(record.Year) if record.Year is not None else None,  # Check for None
            # 'pH': float(record.pH) if record.pH is not None else None,  # Convert to float, and check for None
            # 'Ammonia': float(record.Ammonia) if record.Ammonia is not None else None,  # Convert to float, and check for None
            # 'DO': float(record.DO) if record.DO is not None else None,  # Convert to float, and check for None
            # 'Nitrate': float(record.Nitrate) if record.Nitrate is not None else None,  # Convert to float, and check for None
            # 'Phosphate': float(record.Phosphate) if record.Phosphate is not None else None  # Convert to float, and check for None
            'Year': int(record.Year) if record.Year.strip() else None,  # Check for empty string
            'pH': float(record.pH) if record.pH.strip() else None,  # Convert to float, and check for empty string
            'Ammonia': float(record.Ammonia) if record.Ammonia.strip() else None,  # Convert to float, and check for empty string
            'DO': float(record.DO) if record.DO.strip() else None,  # Convert to float, and check for empty string
            'Nitrate': float(record.Nitrate) if record.Nitrate.strip() else None,  # Convert to float, and check for empty string
            'Phosphate': float(record.Phosphate) if record.Phosphate.strip() else None,  # Convert to float, and check for empty string
            'Time': record.Time,
            'WeatherCondition': record.WeatherCondition,
            'WindDirection': record.WindDirection,
            'ColorOfWater': record.ColorOfWater,
            'AirTemperature': float(record.AirTemperature.strip()) if record.AirTemperature.strip() else None,
            'WaterTransparency': float(record.WaterTransparency.strip()) if record.WaterTransparency.strip() else None,
            'WaterTemperature': float(record.WaterTemperature.strip()) if record.WaterTemperature.strip() else None
    
        }

        db_data_dict.append(record_dict)

    # Convert the list of dictionaries to a Pandas DataFrame
    db_data_df = pd.DataFrame(db_data_dict)
    db_data_df = calculate_wqi(db_data_df)
    river_df = pd.read_csv('taaldata.csv', encoding='ISO-8859-1')  # Load data from CSV file
    river_df = lake_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    river_df = calculate_wqi(river_df)  # Calculate Water Quality Index (WQI)
    numerical_columns = ['pH', 'Ammonia', 'DO', 'Nitrate', 'Phosphate']
    for col in numerical_columns:
        river_df[col] = river_df[col].astype(float)
    
    for data in [db_data_df, river_df]:
        for col in numerical_columns:
            data[col] = data[col].apply(lambda x: None if pd.isna(x) else x)

    # Combine the data from the database and CSV
    combined_data_df = pd.concat([db_data_df, river_df], ignore_index=True)

    print("here db")
    print(combined_data_df)
  
    # Convert the combined data DataFrame to JSON format
    combined_data_json = combined_data_df.to_json(orient='records')
    
    # Convert month names to numerical representations
    #river_df['Month'] = river_df['Month'].map(month_to_number)
    
    #river_df['Year'] = river_df['Year'].astype(int)  # Convert 'Year' column to integers
    combined_data_df['Year'] = combined_data_df['Year'].astype(int)  # Convert 'Year' column to integers
    print('yes')
    print(combined_data_df)
    # Filter data from river_df based on the selected month and year
    filtered_river_df = combined_data_df[(combined_data_df['Month'] == selected_month_number) & (combined_data_df['Year'] == selected_year)]
    
        # Prepare the data to display pH values of the selected barangays
    filtered_data = filtered_river_df[filtered_river_df['Barangay'].isin(selected_barangays_list)]
    
    talisay_count=0
    laurel_count = 0
    Agoncillo_count = 0
    mnk_count = 0
    tanauan_count =0
    brgy_col =[]
    for index, row in filtered_data.iterrows():
        if row['Barangay'] in ['Sampaloc', 'Quiling', 'San Isidro']:
            talisay_count += 1
            brgy_col.append(row['Barangay'])
        if row['Barangay'] in ['Buso-Buso', 'Leviste']:
            laurel_count += 1
            brgy_col.append(row['Barangay'])
        if row['Barangay'] in ['Bañaga', 'Manalaw']:
            Agoncillo_count += 1
            brgy_col.append(row['Barangay'])
        if row['Barangay'] in ['Mataas na Kahoy', 'Nangkaan']:
            mnk_count += 1
            brgy_col.append(row['Barangay'])
        if row['Barangay'] in ['Tanauan']:
            tanauan_count += 1
            brgy_col.append(row['Barangay'])

    ph_values = filtered_data[['Barangay', 'pH']]
    ammonia_values = filtered_data[['Barangay', 'Ammonia']]
    do_values = filtered_data[['Barangay', 'DO']]
    nitrate_values = filtered_data[['Barangay', 'Nitrate']]
    phosphate_values = filtered_data[['Barangay', 'Phosphate']]      
    
    if filtered_river_df.empty:  # Check if the filtered data is empty
        return "No data found"
    # else:
    #     filtered_river_data = WaterQuality.query.filter(
    #         WaterQuality.Month == selected_month_number,  # Use the actual values from your database
    #         WaterQuality.Year == int(selected_year)
    #     ).all()
        
    #     if not filtered_river_data:  # Check if the filtered data from the database is empty
    #         return "No data found"

    #     filtered_river_df = pd.DataFrame([{
    #         'stationid': row.stationid,
    #         'Barangay': row.Barangay,
    #         'Month': month_to_number[row.Month],
    #         'Year': float(row.Year),
    #         'pH': float(row.pH),
    #         'Ammonia': float(row.Ammonia),
    #         'DO': float(row.DO),
    #         'Nitrate': float(row.Nitrate),
    #         'Phosphate': float(row.Phosphate)
    #     } for row in filtered_river_data])
        
    #     filtered_river_df = calculate_wqi(filtered_river_df)  
        
    #     filtered_data = river_df[river_df['Barangay'].isin(selected_barangays_list)]
    #     ph_values = filtered_data[['Barangay', 'pH']]
    #     print(ph_values)
    #     ammonia_values = filtered_data[['Barangay', 'Ammonia']]
    #     do_values = filtered_data[['Barangay', 'DO']]
    #     nitrate_values = filtered_data[['Barangay', 'Nitrate']]
    #     phosphate_values = filtered_data[['Barangay', 'Phosphate']]      
    
    #filtered_river_df = pd.concat([data_from_db, filtered_river_df], ignore_index=True)

    
    
# Check if the concatenated DataFrame is empty
    # Define sample data
     # Create a PDF document
     #-------- page1
    phyparam = pd.read_csv('physicalParam.csv', encoding='ISO-8859-1')  # Load data from CSV file
    phyparam = phyparam.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    
    # Convert month names to numerical representations
    #river_df['Month'] = river_df['Month'].map(month_to_number)
    phyparam['Year'] = phyparam['Year'].astype(int)  # Convert 'Year' column to integers
    phyparam = phyparam[(phyparam['Month'] == selected_month_number) & (river_df['Year'] == selected_year)]
    
    
    
    phyparam_filtered = phyparam[phyparam['Barangay'].isin(selected_barangays_list)]
    
    phyparam_filtered = pd.merge(db_data_df, phyparam_filtered, on='Barangay', how='outer')
    print("here db")
    print(phyparam_filtered)
    time_values = phyparam_filtered[['Barangay', 'Time (AM)']]
    weather_values = phyparam_filtered[['Barangay', 'Weather Condition']]
    wind_values = phyparam_filtered[['Barangay', 'Wind Direction']]
    color_values = phyparam_filtered[['Barangay', 'Color of Water (apparent)']]
    air_values = phyparam_filtered[['Barangay', 'Air Temperature (0C)']]
    water_transparent_values = phyparam_filtered[['Barangay', 'Water Transparency (m)']]
    depth_values = phyparam_filtered[['Barangay', 'Depth, m']]
    water_temp_values = phyparam_filtered[['Barangay', 'Water Temperature (0C)']]
    hardness_values = phyparam_filtered[['Barangay', 'Hardness (mg/L)']]
    
    
    pdf = FPDF(orientation='L', unit='mm', format='A4')  # Landscape orientation
    pdf.add_page()
    
    pdf.set_font("Arial", size=12)

    create_header(pdf, selected_month_number, selected_year)
    col_width = 39.57
    brgy_column = (col_width * 5) / (talisay_count+laurel_count+Agoncillo_count+mnk_count+tanauan_count)
    header = ["Monitoring Station"] + [param for param in [ "DO", "pH", "Phosphate", "Ammonia", "Nitrate"]]
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(col_width, 3, "", 'LTR', 0, 'C')
    
    
    for i in range(len(selected_stations)):
        station = selected_stations[i]
        if(station == 'Talisay'):
            pdf.cell(brgy_column*talisay_count, 3, "", 'LTR', 0, 'C')
        if(station == 'Laurel'):
            pdf.cell(brgy_column*laurel_count, 3, "", 'LTR', 0, 'C')
        if(station == 'Agoncillo'):
            pdf.cell(brgy_column*Agoncillo_count, 3, "", 'LTR', 0, 'C')
        if(station == 'Mataas na Kahoy'):
            pdf.cell(brgy_column*mnk_count, 3, "", 'LTR', 0, 'C')
        if(station == 'Tanauan'):
            pdf.cell(brgy_column*tanauan_count, 3, "", 'LTR', 0, 'C')
    
    pdf.cell(col_width, 3, "", 'LTR', 1, 'C')
    pdf.cell(col_width, 3, "", 'LR', 0, 'C')
    
    for i in range(len(selected_stations)):
        station = selected_stations[i]
        if(station == 'Talisay'):
            pdf.cell(brgy_column*talisay_count, 3, "TALISAY", 'LR', 0, 'C')
        if(station == 'Laurel'):
            pdf.cell(brgy_column*laurel_count, 3, "LAUREL", 'LR', 0, 'C')
        if(station == 'Agoncillo'):
            pdf.cell(brgy_column*Agoncillo_count, 3, "AGONCILLO", 'LR', 0, 'C')
        if(station == 'Mataas na Kahoy'):
            if(brgy_count == 7):
                pdf.set_font('Arial', 'B', 9.5)
            elif(brgy_count > 7):
                pdf.set_font('Arial', 'B', 6)
            pdf.cell(brgy_column*mnk_count, 3, "MATAAS NA KAHOY", 'LR', 0, 'C')
        if(station == 'Tanauan'):
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(brgy_column*tanauan_count, 3, "TANAUAN", 'LR', 0, 'C')
        
        
    
    pdf.cell(col_width, 3, "Standard Levels", 'LR', 1, 'C')
    pdf.cell(col_width, 3, "", 'LR', 0, 'C')
    
    for i in range(len(selected_stations)):
        station = selected_stations[i]
        if(station == 'Talisay'):
            pdf.cell(brgy_column*talisay_count, 3, "", 'LBR', 0, 'C')
        if(station == 'Laurel'):
            pdf.cell(brgy_column*laurel_count, 3, "", 'LBR', 0, 'C')
        if(station == 'Agoncillo'):
            pdf.cell(brgy_column*Agoncillo_count, 3, "", 'LBR', 0, 'C')
        if(station == 'Mataas na Kahoy'):
            pdf.cell(brgy_column*mnk_count, 3, "", 'LBR', 0, 'C')
        if(station == 'Tanauan'):
            pdf.cell(brgy_column*tanauan_count, 3, "", 'LBR', 0, 'C')
    
    pdf.cell(col_width, 3, "for Class C Waters", 'LR', 1, 'C')
    
    pdf.cell(col_width, 3, "CHEMICAL", 'LR', 0, 'C')
    
    for i in range(len(brgy_col)):
        pdf.cell(brgy_column, 3, '', 'LR', 0, 'C')
    
    pdf.set_font('Arial', '', 10)
    pdf.cell(col_width, 3, "(* DAO 2016-08,", 'LR', 1, 'C')
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(col_width, 3, "PARAMETERS", 'LR', 0, 'C')
    
    for i in range(len(brgy_col)):
        if(brgy_col[i] == 'Mataas na Kahoy' and brgy_count == 7):
            pdf.set_font('Arial', '', 9)
        elif(brgy_col[i] == 'Mataas na Kahoy' and brgy_count > 7):
            pdf.set_font('Arial', '', 7)
        else:
            pdf.set_font('Arial', '', 10)
        pdf.cell(brgy_column, 3, brgy_col[i], 'LR', 0, 'C')
    
    pdf.set_font('Arial', '', 10)
    pdf.cell(col_width, 3, "** DAO 2021-19,", 'LR', 1, 'C')
    pdf.cell(col_width, 3, "", 'LR', 0, 'C')
    for i in range(len(brgy_col)):
        pdf.cell(brgy_column, 3,'', 'LR', 0, 'C')
    pdf.cell(col_width, 3, "*** Abowei, 2010)", 'LBR', 1, 'C')
    
    param_list = ['Time (AM)', 'Weather Condition',  'Wind Direction', 'Color of Water (apparent)','Air Temperature (0C)','Water Transparency (m)','Water Temperature (0C)']
    for i in range(len(param_list)):
        pdf.set_font("Times", size=10)  # Set font size to 8
        pdf.cell(39.57, 8, param_list[i], border=1, ln=False, align='C')
        
        if i == 0:
            for index, row in time_values.iterrows():
                value = row['Time (AM)'] if pd.notnull(row['Time (AM)']) else 'ND'
                pdf.cell(brgy_column, 8, str(value), border=1, ln=False, align='C')
            pdf.cell(col_width, 8, '', 'LTR', ln=False, align='C') 
        elif i == 1:
            for index, row in weather_values.iterrows():
                value = row['Weather Condition'] if pd.notnull(row['Weather Condition']) else 'ND'
                pdf.cell(brgy_column, 8, str(value), border=1, ln=False, align='C')
            pdf.cell(col_width, 8, '', 'LR', ln=False, align='C')
        elif i == 2:
            for index, row in wind_values.iterrows():
                value = row['Wind Direction'] if pd.notnull(row['Wind Direction']) else 'ND'
                pdf.cell(brgy_column, 8, str(value), border=1, ln=False, align='C')
            pdf.cell(col_width, 8, '', 'LR', ln=False, align='C')
        elif i == 3:
            for index, row in color_values.iterrows():
                value = row['Color of Water (apparent)'] if pd.notnull(row['Color of Water (apparent)']) else 'ND'
                pdf.cell(brgy_column, 8, str(value), border=1, ln=False, align='C')
            pdf.cell(col_width, 8, '', 'LR', ln=False, align='C')
        elif i == 4:
            for index, row in air_values.iterrows():
                value = row['Air Temperature (0C)'] if pd.notnull(row['Air Temperature (0C)']) else 'ND'
                pdf.cell(brgy_column, 8, str(value), border=1, ln=False, align='C')
            pdf.cell(col_width, 8, '', 'LR', ln=False, align='C')
        elif i == 5:
            for index, row in water_transparent_values.iterrows():
                value = row['Water Transparency (m)'] if pd.notnull(row['Water Transparency (m)']) else 'ND'
                pdf.cell(brgy_column, 8, str(value), border=1, ln=False, align='C')
            pdf.cell(col_width, 8, '', 'LR', ln=False, align='C')
        # if(i == 6):
        #     for index, row in depth_values.iterrows():
        #         pdf.cell(brgy_column,8, str(row['Depth, m']),border=1, ln=False, align='C')
        #     pdf.cell(col_width,8,'','LR', ln=False, align='C')
        elif i == 6:
            for index, row in water_temp_values.iterrows():
                value = row['Water Temperature (0C)'] if pd.notnull(row['Water Temperature (0C)']) else 'ND'
                pdf.cell(brgy_column, 8, str(value), border=1, ln=False, align='C')
            pdf.cell(col_width, 8, '', 'LBR', ln=False, align='C')
        # if(i == 8):
        #     for index, row in hardness_values.iterrows():
        #         pdf.cell(brgy_column,8, str(row['Hardness (mg/L)']),border=1, ln=False, align='C')
        #     pdf.cell(col_width,8,'','LBR', ln=False, align='C')

    
        pdf.ln()

    pdf.set_font("Times", 'B', 7)
    pdf.cell(0, 5, "LEGEND: ", '', 1, 'L')
    
    #----2nd page----------------------------------------------------------------------------------------------------------------------------------------
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    create_header(pdf, selected_month_number, selected_year)

    
   # pdf.cell(280, 10, f"River Report for Quarter {selected_quarter}", ln=True, align='C')  # Adjust width for landscape
    #pdf.ln(10)
    col_width = 39.57
    brgy_column = (col_width * 5) / (talisay_count+laurel_count+Agoncillo_count+mnk_count+tanauan_count)
    header = ["Monitoring Station"] + [param for param in [ "DO", "pH", "Phosphate", "Ammonia", "Nitrate"]]
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(col_width, 3, "", 'LTR', 0, 'C')
    
    
    for i in range(len(selected_stations)):
        station = selected_stations[i]
        if(station == 'Talisay'):
            pdf.cell(brgy_column*talisay_count, 3, "", 'LTR', 0, 'C')
        if(station == 'Laurel'):
            pdf.cell(brgy_column*laurel_count, 3, "", 'LTR', 0, 'C')
        if(station == 'Agoncillo'):
            pdf.cell(brgy_column*Agoncillo_count, 3, "", 'LTR', 0, 'C')
        if(station == 'Mataas na Kahoy'):
            pdf.cell(brgy_column*mnk_count, 3, "", 'LTR', 0, 'C')
        if(station == 'Tanauan'):
            pdf.cell(brgy_column*tanauan_count, 3, "", 'LTR', 0, 'C')
    
    pdf.cell(col_width, 3, "", 'LTR', 1, 'C')
    pdf.cell(col_width, 3, "", 'LR', 0, 'C')
    
    for i in range(len(selected_stations)):
        station = selected_stations[i]
        if(station == 'Talisay'):
            pdf.cell(brgy_column*talisay_count, 3, "TALISAY", 'LR', 0, 'C')
        if(station == 'Laurel'):
            pdf.cell(brgy_column*laurel_count, 3, "LAUREL", 'LR', 0, 'C')
        if(station == 'Agoncillo'):
            pdf.cell(brgy_column*Agoncillo_count, 3, "AGONCILLO", 'LR', 0, 'C')
        if(station == 'Mataas na Kahoy'):
            if(brgy_count == 7):
                pdf.set_font('Arial', 'B', 9.5)
            elif(brgy_count > 7):
                pdf.set_font('Arial', 'B', 6)
            pdf.cell(brgy_column*mnk_count, 3, "MATAAS NA KAHOY", 'LR', 0, 'C')
        if(station == 'Tanauan'):
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(brgy_column*tanauan_count, 3, "TANAUAN", 'LR', 0, 'C')
        
        
    
    pdf.cell(col_width, 3, "Standard Levels", 'LR', 1, 'C')
    pdf.cell(col_width, 3, "", 'LR', 0, 'C')
    
    for i in range(len(selected_stations)):
        station = selected_stations[i]
        if(station == 'Talisay'):
            pdf.cell(brgy_column*talisay_count, 3, "", 'LBR', 0, 'C')
        if(station == 'Laurel'):
            pdf.cell(brgy_column*laurel_count, 3, "", 'LBR', 0, 'C')
        if(station == 'Agoncillo'):
            pdf.cell(brgy_column*Agoncillo_count, 3, "", 'LBR', 0, 'C')
        if(station == 'Mataas na Kahoy'):
            pdf.cell(brgy_column*mnk_count, 3, "", 'LBR', 0, 'C')
        if(station == 'Tanauan'):
            pdf.cell(brgy_column*tanauan_count, 3, "", 'LBR', 0, 'C')
    
    pdf.cell(col_width, 3, "for Class C Waters", 'LR', 1, 'C')
    
    pdf.cell(col_width, 3, "CHEMICAL", 'LR', 0, 'C')
    
    for i in range(len(brgy_col)):
        pdf.cell(brgy_column, 3, '', 'LR', 0, 'C')
    
    pdf.set_font('Arial', '', 10)
    pdf.cell(col_width, 3, "(* DAO 2016-08,", 'LR', 1, 'C')
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(col_width, 3, "PARAMETERS", 'LR', 0, 'C')
    
    for i in range(len(brgy_col)):
        if(brgy_col[i] == 'Mataas na Kahoy' and brgy_count == 7):
            pdf.set_font('Arial', '', 9)
        elif(brgy_col[i] == 'Mataas na Kahoy' and brgy_count > 7):
            pdf.set_font('Arial', '', 7)
        else:
            pdf.set_font('Arial', 'B', 10)
        pdf.cell(brgy_column, 3, brgy_col[i], 'LR', 0, 'C')
    
    pdf.set_font('Arial', '', 10)
    pdf.cell(col_width, 3, "** DAO 2021-19,", 'LR', 1, 'C')
    pdf.cell(col_width, 3, "", 'LR', 0, 'C')
    for i in range(len(brgy_col)):
        pdf.cell(brgy_column, 3,'', 'LR', 0, 'C')
    pdf.cell(col_width, 3, "*** Abowei, 2010)", 'LBR', 1, 'C')
    # pdf.cell(col_width, 9, "", 'LR', 0, 'C')
        
    m1=""
    m2=""
    m3=""
    # if selected_quarter == "Q1":
    #     m1 = "Jan"
    #     m2 = "Feb"
    #     m3 = "March"
    # elif selected_quarter == "Q2":
    #     m1 = "April"
    #     m2 = "May"
    #     m3 = "June"
    # elif selected_quarter == "Q3":
    #     m1 = "July"
    #     m2 = "Aug"
    #     m3 = "Sept"
    # else:
    #     m1 = "Oct"
    #     m2 = "Nov"
    #     m3 = "Dec"
        
    # pdf.cell(13.19, 10, m1, 'LBR', 0, 'C')
    # pdf.cell(13.19, 10, m2, 'LBR', 0, 'C')
    # pdf.cell(13.19, 10, m3, 'LBR', 0, 'C')   
    # pdf.cell(13.19, 10, m1, 'LBR', 0, 'C')
    # pdf.cell(13.19, 10, m2, 'LBR', 0, 'C')
    # pdf.cell(13.19, 10, m2, 'LBR', 0, 'C')
    # pdf.cell(13.19, 10, m3, 'LBR', 0, 'C')   
    # pdf.cell(13.19, 10, m1, 'LBR', 0, 'C')
    # pdf.cell(13.19, 10, m2, 'LBR', 0, 'C')
    # pdf.cell(13.19, 10, m3, 'LBR', 0, 'C')   
    # pdf.cell(13.19, 10, m2, 'LBR', 0, 'C')
    # pdf.cell(13.19, 10, m3, 'LBR', 0, 'C')   
    # pdf.cell((col_width*5)/3, 9, m1, 'LBR', 0, 'C')
    # pdf.cell((col_width*5)/3, 9, m2, 'LBR', 0, 'C')
    # pdf.cell((col_width*5)/3, 9, m3, 'LBR', 0, 'C')
    
    # pdf.cell(col_width, 9, "", 'LBR', 1, 'C')
    
    
    # for h in header:
    #     pdf.cell(col_width, 10, h, border=1, ln=False, align='C')
    # pdf.ln()
    # # Add the first blank row with borders
    # for _ in range(2):  # Adding two blank rows
    #     for _ in range(len(header)):
    #         pdf.cell(col_width, 10, "", border=1)  # Empty cell with borders
    #     pdf.ln()  # Move to the next line
        
# Iterate through each Barangay first
    param_list = ['pH', 'Amonia-N(mg/L)', 'Phospate(mg/L)', 'Nitrate-N(mg/L)', 'Dissolve Oxygen(mg/L)']
    for i in range(len(param_list)):
        pdf.set_font("Times", size=10)  # Set font size to 8
        pdf.cell(39.57, 8, param_list[i], border=1, ln=False, align='C')
        
        if(i == 0):
            for index, row in ph_values.iterrows():
                ph_value = float(row['pH'])
                if(ph_value < 6.50 or ph_value > 9.00):
                    pdf.set_text_color(255,0,0)
                else:
                    pdf.set_text_color(0,0,0)

                if str(row['pH']) != '':
                    pdf.cell(brgy_column,8,str(row['pH']),border=1, ln=False, align='C')
                else:
                    pdf.cell(brgy_column,8,'ND',border=1, ln=False, align='C')
            pdf.set_text_color(0,0,0)
            pdf.cell(col_width,8,'6.50 - 9.00*',border=1, ln=False, align='C') 
        if(i == 1):
            for index, row in ammonia_values.iterrows():
                ammonia_values = float(row['Ammonia'])
                if(ammonia_values < 0.06):
                    pdf.set_text_color(255,0,0)
                else:
                    pdf.set_text_color(0,0,0)
                    
                if str(row['Ammonia']) != '':
                    pdf.cell(brgy_column,8,str(row['Ammonia']),border=1, ln=False, align='C')
                else:
                    pdf.cell(brgy_column,8,'ND',border=1, ln=False, align='C')
            pdf.set_text_color(0,0,0)
            pdf.cell(col_width,8,'<0.06**',border=1, ln=False, align='C')
        if(i == 2):
            for index, row in phosphate_values.iterrows():
                phosphate_values = float(row['Phosphate'])
                if(phosphate_values < 0.025):
                    pdf.set_text_color(255,0,0)
                else:
                    pdf.set_text_color(0,0,0)
                if str(row['Phosphate']) != '':
                    pdf.cell(brgy_column,8,str(row['Phosphate']),border=1, ln=False, align='C')
                else:
                    pdf.cell(brgy_column,8,'ND',border=1, ln=False, align='C')
            pdf.set_text_color(0,0,0)
            pdf.cell(col_width,8,'<0.025**',border=1, ln=False, align='C')
        if(i == 3):
            for index, row in nitrate_values.iterrows():
                nitrate_values = float(row['Nitrate'])
                if(nitrate_values < 0.06):
                    pdf.set_text_color(255,0,0)
                else:
                    pdf.set_text_color(0,0,0)
                if str(row['Nitrate']) != '':
                    pdf.cell(brgy_column,8,str(row['Nitrate']),border=1, ln=False, align='C')
                else:
                    pdf.cell(brgy_column,8,'ND',border=1, ln=False, align='C')
            pdf.set_text_color(0,0,0)
            pdf.cell(col_width,8,'<7.00*',border=1, ln=False, align='C')
        if(i == 4):
            for index, row in do_values.iterrows():
                do_values = float(row['DO'])
                if(do_values < 0.06):
                    pdf.set_text_color(255,0,0)
                else:
                    pdf.set_text_color(0,0,0)

                if str(row['DO']) != '':
                    pdf.cell(brgy_column,8, f"S - {row['DO']}",border=1, ln=False, align='C')
                else:
                    pdf.cell(brgy_column,8,'ND',border=1, ln=False, align='C')
                    
            pdf.set_text_color(0,0,0)
            pdf.cell(col_width,8,'>5.00*',border=1, ln=False, align='C')

        pdf.ln()

# ... rest of the code
    pdf.set_font("Times", 'B', 7)
    pdf.cell(0, 5, "METHODOLOGY: Water Quality Multi-Parameter Checker (YSI EXO). La Motte Smart 3 Colorimeter, EXTECH W q 500", '', 1, 'L')
    
    pdf.ln(10)
    pdf.set_font("Times", 'B', 10)
    pdf.cell(130, 5, "Prepared by:", '', 0, 'L')
    pdf.cell(0, 5, "Approved and checked by:", '', 1, 'L')
    
    pdf.ln(5)
    pdf.set_font("Times", 'B', 12)
    pdf.cell(30, 5, "", '', 0, 'L')
    pdf.cell(60, 5, "SHAIRA C. JAVIER, RCh.", '', 0, 'C')
    pdf.cell(20, 5, "", '', 0, 'L')
    pdf.cell(0, 5, "NENITA S. KAWIT", '', 1, 'C')
    pdf.set_font("Times", '', 12)
    pdf.cell(30, 5, "", '', 0, 'L')
    pdf.cell(60, 5, "License No. 0014902", '', 0, 'C')
    pdf.cell(20, 5, "", '', 0, 'L')
    pdf.cell(0, 5, "Center Chief-BIFTOS", '', 1, 'C')
    


    # Output the PDF to response
    response = make_response(pdf.output(dest='S').encode('latin1'))
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=river_report_{selected_month_number}.pdf'
    return response

# ... (other code)


from sklearn.ensemble import (
    RandomForestClassifier, VotingClassifier
)
from sklearn.model_selection import (
    train_test_split, RepeatedStratifiedKFold, cross_val_score
)
from sklearn.metrics import accuracy_score, classification_report
from sklearn.preprocessing import LabelEncoder
from xgboost import XGBClassifier
from lightgbm import LGBMClassifier
from catboost import CatBoostClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from time import time
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import confusion_matrix
@app.route('/result2')
def result2():
    df = pd.read_csv("taaldata.csv", encoding='ISO-8859-1')

    # Handling Missing Values
    df['DO'] = df['DO'].fillna(df['DO'].mean())
    df['pH'] = df['pH'].fillna(df['pH'].mean())
    df['Phosphate'] = df['Phosphate'].fillna(df['Phosphate'].mean())
    df['Ammonia'] = df['Ammonia'].fillna(df['Ammonia'].mean())
    df['Nitrate'] = df['Nitrate'].fillna(df['Nitrate'].mean())

    for col in ['DO','pH','Phosphate','Ammonia','Nitrate']:
        df[col] = (df[col] - df[col].mean()) / df[col].std()

    data_df = calculate_wqi(df)
    
    df.drop(['stationid','Barangay','Month','Year'],axis=1,inplace=True)

    # x=df.iloc[:,1:4].values 
    # y=df.iloc[:,-1].values.reshape(-1,1) #wqc column
    
    # Features and target variable for XGBoost
    X = df[['DO', 'pH', 'Phosphate', 'Ammonia', 'Nitrate']]
    y = df['wqc']

    # Feature Engineering: Applying PCA for dimensionality reduction
    pca = PCA(n_components=3)
    x_pca = pca.fit_transform(X)

    # Normalizing and Scaling: Standardize features
    scaler = StandardScaler()
    x_scaled = scaler.fit_transform(x_pca)

    # Split data into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    #randomforest -------------------------------------------------------------------
    rf = RandomForestClassifier()

    param_grid_rf = {
        'n_estimators': [100, 200, 300,400,500],
        'max_depth': [None, 10, 20, 30,40,50],
        'min_samples_split': [2, 5, 10],
        'min_samples_leaf': [1, 2, 4],
        'bootstrap': [True, False],
    }

    # Perform Random Search for Random Forest/# Initialize the RandomizedSearchCV object
    random_search_rf = RandomizedSearchCV(estimator=rf, param_distributions=param_grid_rf, n_iter=5, cv=5, scoring='accuracy', random_state=100, n_jobs=-1)
    random_search_rf.fit(X_train, np.ravel(y_train))
    
    best_rf_estimator = random_search_rf.best_estimator_
    best_rf_score = random_search_rf.best_score_

    rf = RandomForestClassifier(max_depth=6, n_estimators=100)
    #rf1 = RandomForestClassifier(rf_rs.best_params_)
    
    rf.fit(X_train, np.ravel(y_train))

    pred_rf = rf.predict(X_test)

    rf_score = cross_val_score(rf,X,np.ravel(y),cv=5)
    rf_score_mean = rf_score.mean()
    
    acc_rf = accuracy_score(y_test, pred_rf)
    
    target_names = ['Excellent', 'Good', 'Poor', 'Very Poor', 'Unsuitable']
    cr_rf = classification_report(y_test,pred_rf)
    print(f"RandomForest Classification Report = \n {cr_rf}")

    #Bagging - randomforest base estimator-----------------------------------------
    baggingrf = BaggingClassifier()

    param_grid_bagging = {
        'base_estimator__max_depth' : [1, 2, 3, 4, 5],
        'max_samples' : [0.05, 0.1, 0.2, 0.5]
    }

    bagging_rs=RandomizedSearchCV(BaggingClassifier(RandomForestClassifier(), n_estimators = 100, max_features = 0.5), param_grid_bagging)

    baggingrf.fit(X_train, np.ravel(y_train))

    pred_bagging = baggingrf.predict(X_test)

    bagging_score = cross_val_score(rf,X,np.ravel(y),cv=5)
    bagging_score_mean = bagging_score.mean()
    
    acc_bagging = accuracy_score(y_test, pred_bagging)
  
    target_names = ['Excellent', 'Good', 'Poor', 'Very Poor', 'Unsuitable']
    cr_bagging = classification_report(y_test,pred_bagging)
    print(f"RandomForest Classification Report = \n {cr_bagging}")

    #SVM ------------------------------------------------------------
    #--------------------Support Vector Machine ----------
    svmparams={
        'C': [1, 10, 100],
        'gamma': [1, 0.1, 0.01, 0.001, 0.0001],
        'kernel': ['rbf','linear']
    }

    svm = SVC()
    svm_rs=RandomizedSearchCV(svm,param_distributions=svmparams,n_iter=5,scoring='accuracy',n_jobs=-1,cv=5)
    svm_rs.fit(X_train, np.ravel(y_train))

    best_svm_est = svm_rs.best_estimator_
    
    best_svm_score = svm_rs.best_score_
    
    svm1 = SVC(C=10, gamma=0.1, probability=True, kernel='linear')
    #svm1 = SVC(C=100, gamma=0.001, kernel='linear')
    #svm1 = SVC(svm_rs.best_estimator_)
    svm1.fit(X_train, np.ravel(y_train))

    pred_svm = svm1.predict(X_test)

    svm_score = cross_val_score(svm1,X,np.ravel(y),cv=5)
    svm_score_mean = svm_score.mean()

    acc_svm = accuracy_score(y_test, pred_svm)

    target_names = ['Excellent', 'Good', 'Poor', 'Very Poor', 'Unsuitable']
    cr_svm = classification_report(y_test,pred_svm)
    print(f"RandomForest Classification Report = \n {cr_svm}")

    # naeve bayes -----------------------------------
    #DT
    dtparams={
        'criterion' : ['gini', 'entropy'],
        'max_depth' : [3, 5, 7, 10],
        'splitter' : ['best', 'random'],
        'min_samples_leaf' : [1, 2, 3, 5, 7],
    #    'max_features' : ['auto', 'sqrt', 'log2']
    }

    dt = DecisionTreeClassifier()

    dt_rs=RandomizedSearchCV(dt,param_distributions=dtparams,n_iter=5,scoring='accuracy',n_jobs=-1,cv=5)
    dt_rs.fit(X_train, np.ravel(y_train))

    dt_bestest=dt_rs.best_estimator_
    dt_rs.best_params_
    dt_bestparams=dt_rs.best_params_
    dt_bestscore=dt_rs.best_score_

    
    #dt1 = DecisionTreeClassifier(splitter='best',min_samples_leaf=1,max_features='log2',max_depth=3,criterion='gini')
    dt1 = DecisionTreeClassifier(splitter='best',min_samples_leaf=1,max_depth=3,criterion='gini')
    #dt1 = DecisionTreeClassifier(dt_rs.best_estimator_)
    dt1.fit(X_train, np.ravel(y_train))
    pred_dt1 = dt1.predict(X_test)

    dt_score = cross_val_score(dt1,X,np.ravel(y),cv=5)
    dt_score_mean = dt_score.mean()

    acc_dt = accuracy_score(y_test, pred_dt1)

    target_names = ['Excellent', 'Good', 'Poor', 'Very Poor', 'Unsuitable']
    cr_dt = classification_report(y_test,pred_dt1)
    print(f"Decision Tree Classification Report = \n {cr_dt}")

    # nb = MultinomialNB()
    
    # nbparams = {'alpha': [0.1, 0.5, 1.0, 2.0, 5.0, 10.0]}

    # #nb_rs=RandomizedSearchCV(nb,param_distributions=nbparams,n_iter=5,scoring='accuracy',cv=5)
    # nb_rs=RandomizedSearchCV(nb,param_distributions=nbparams,n_iter=5,scoring='accuracy',cv=5)
    # nb_rs.fit(X_train, np.ravel(y_train))

    # best_nb_est = nb_rs.best_estimator_
    
    # best_nb_score = nb_rs.best_score_
    
    # best_alpha = nb_rs.best_params_['alpha']

    # nb1 = MultinomialNB(alpha=best_alpha)
    # nb1.fit(X_train, np.ravel(y_train))

    # pred_nb = nb1.predict(X_test)

    # nb_score = cross_val_score(nb1,X,np.ravel(y),cv=5)
    # nb_score_mean = nb_score.mean()

    # acc_nb = accuracy_score(y_test, pred_nb)

    # #target_names = ['Excellent', 'Good', 'Poor', 'Very Poor', 'Unsuitable']
    # #cr_nb = classification_report(y_test,pred_nb, target_names=target_names)
    # cr_nb = classification_report(y_test,pred_nb)
    # print(f"RandomForest Classification Report = \n {cr_nb}")
    #-------------------------------------------------------------------

    user_id = session.get('user_id')
    fname = session.get('fname')
    lname = session.get('lname')
    profile_image = session.get('profile_image')
    userType = session.get('userType')

    return render_template('result2.html', 
                        best_rf_estimator=best_rf_estimator,
                        best_rf_score=best_rf_score,
                        rf_score_mean=rf_score_mean,
                        acc_rf=acc_rf,
                        cr_rf=cr_rf,

                        bagging_score_mean=bagging_score_mean,
                        acc_bagging=acc_bagging,
                        cr_bagging=cr_bagging,

                        best_svm_est=best_svm_est,
                        best_svm_score=best_svm_score,
                        svm_score_mean=svm_score_mean,
                        acc_svm=acc_svm,
                        cr_svm=cr_svm,

                        acc_dt=acc_dt,
                        dt_bestscore=dt_bestscore,
                        cr_dt=cr_dt,
                        dt_bestest=dt_bestest,
                        dt_score_mean=dt_score_mean,

                        user_id=user_id, 
                        fname=fname, 
                        lname=lname, 
                        profile_image=profile_image,
                        userType=userType)

if __name__ == '__main__':
    lake_df = pd.read_csv('taaldata.csv' , encoding='ISO-8859-1')
    lake_df = calculate_and_map_quarters(lake_df)  # Calculate WQI and map quarters
    lake_csv_data_with_wqi_json = lake_df.to_json(orient='records')  # Convert DataFrame to JSON string

    river_df = pd.read_csv('taaldata.csv' , encoding='ISO-8859-1')
    river_df = calculate_and_map_quarters(river_df)  # Calculate WQI and map quarters
    river_csv_data_with_wqi_json = river_df.to_json(orient='records')
    app.run(debug=True, port=2200,threaded=True)
